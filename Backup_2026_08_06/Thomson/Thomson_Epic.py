
#----------###################################----------------------NEW PREMIUM CONDITION WITH LOB SPECIFICATION--------------------######################################--------------------------


# import asyncio
# from email import policy
# from playwright.async_api import async_playwright
# from configparser import ConfigParser
# from datetime import datetime, timedelta
# import pandas as pd
# import os
# import re
# import extract_msg
# from openpyxl.utils import get_column_letter
# import traceback

# BASE_FOLDER = r"C:\Users\vijay_m\OneDrive - Exdion Solutions Pvt. Ltd.-70692290\Documents\Project_Job_Creation\Thomson"
# PDF_FOLDER = os.path.join(BASE_FOLDER, "Job_Creation")
# EXCEL_FILE_PATH = os.path.join(BASE_FOLDER, "Thomson_Tracker.xlsx")


# # ================= GET STARTING REFERENCE NUMBER =================

# def get_starting_ref_id():

#     if not os.path.exists(EXCEL_FILE_PATH):
#         return 1

#     df = pd.read_excel(EXCEL_FILE_PATH)

#     if "Reference ID" not in df.columns or df.empty:
#         return 1

#     existing_ids = df["Reference ID"].dropna()

#     numbers = (
#         existing_ids
#         .str.extract(r'REF(\d+)')[0]
#         .dropna()
#         .astype(int)
#     )

#     return numbers.max() + 1 if not numbers.empty else 1



# # LOGIN FUNCTION

# async def login(page, context):

#     config = ConfigParser()
#     config.read(os.path.join(BASE_FOLDER, "Config.ini"))

#     url = config.get("ThomsonEpic", "Link")
#     EnterpriseID = config.get("ThomsonEpic", "Enterprise_ID")
#     username = config.get("ThomsonEpic", "User_ID")
#     password = config.get("ThomsonEpic", "Password")

#     await page.goto(url, timeout=90000)

#     await page.get_by_role("textbox").fill(EnterpriseID)

#     await page.locator("//button[.//span[contains(text(),'Contin')]]").click()

#     async with context.expect_page() as new_page_event:
#         await page.locator("//button[contains(text(),'Login')]").click()

#     login_page = await new_page_event.value

#     await login_page.locator("#usercode").fill(username)
#     await login_page.locator("#password").fill(password)

#     await login_page.locator("//button[.//div[text()='Login']]").click()

#     try:
#         later_btn = login_page.get_by_role("button", name="Later")
#         #login_page.locator("button:has-text('Later')").click()
#         await later_btn.wait_for(timeout=5000)
#         await later_btn.click()
#         print("Later Reset clicked")
#     except:
#         pass
    
#     # login_page.wait_for_timeout(5000)
#     # #login_page.wait_for_event("close")  
#     # page.wait_for_load_state("networkidle")

#     await page.locator("//div[contains(@class,'drop-btn')]").first.click()

#     await page.locator("//span[@class='text' and text()='TSLIN01_PROD']").click()

#     await page.locator("//button[.//span[contains(text(),'Contin')]]").click()

#     try:
#         await page.locator("//button[@data-automation-id='Yes']").click(timeout=3000)
#         print("Popup YES clicked")
#     except:
#         print("Login successful")

#     # await page.pause()

#     return login_page


# #--------------------------------------------It Works if .msg file in last-----------------------------------------------------------------
# import os
# import extract_msg

# def extract_pdf_from_msg(msg_path):
#     try:
#         msg = extract_msg.Message(msg_path)
#         save_folder = os.path.dirname(msg_path)

#         target_keywords = [
#             "qte",
#             "quote",
#             "binder",
#             "proposal",
#             "schedule",
#             "application",
#             "app",
#             "accord"
#         ]

#         # Get existing files in folder
#         existing_files = [f.lower() for f in os.listdir(save_folder)]

#         for att in msg.attachments:
#             filename = att.longFilename or att.shortFilename

#             if not filename:
#                 continue

#             filename_lower = filename.lower()

#             if filename_lower.endswith(".pdf"):
                
#                 matched_keyword = next((k for k in target_keywords if k in filename_lower), None)

#                 if matched_keyword:

#                     # 🔥 Check if ANY existing file already has this keyword
#                     if any(matched_keyword in f for f in existing_files):
#                         print(f"Skipping (already have {matched_keyword}):", filename)
#                         continue

#                     pdf_path = os.path.join(save_folder, filename)

#                     with open(pdf_path, "wb") as f:
#                         f.write(att.data)

#                     print("Extracted:", pdf_path)

#     except Exception as e:
#         print("MSG Extraction Error:", e)


# #-------------------------------------     Premium Scrap Extraction Logic     ---------------------------------------------
        
# from datetime import datetime
# from playwright.async_api import expect

# async def policy_premium(page, lob, policy_no, exp_eff_year=2026, exp_exp_year=2027):

#     rows = page.locator('[data-automation-id*="vlvwPolicies body-row"]')

#     await rows.first.wait_for(timeout=15000)

#     count = await rows.count()
#     print("Row count:", count)

#     for i in range(count):

#         row = rows.nth(i)
#         cells = row.locator('.body-cell .text')

#         try:
#             line = (await cells.nth(0).text_content() or "").strip()
#             effective = (await cells.nth(2).text_content() or "").strip()
#             expiration = (await cells.nth(3).text_content() or "").strip()
#             policy_number = (await cells.nth(4).text_content() or "").strip()

#             #print(f"Checking Row {i}: LOB={line}, Effective={effective}, Expiration={expiration}")

#             print(
#                 f"Checking Row {i}: "
#                 f"LOB={line}, "
#                 f"Policy={policy_number}, "
#                 f"Effective={effective}, "
#                 f"Expiration={expiration}"
#             )

#             if lob.lower() not in line.lower():
#                 continue

#             if policy_no.strip().lower() != policy_number.lower():
#                 continue

#             try:
#                 eff_year = datetime.strptime(effective, "%m/%d/%Y").year
#                 exp_year = datetime.strptime(expiration, "%m/%d/%Y").year
#             except:
#                 continue

#         except Exception as e:
#             print("Row Error:", e)
#             continue

#         if eff_year == exp_eff_year and exp_year == exp_exp_year:

#             print(f"Match found at row {i}, opening...")

#             await row.dblclick()

#             # =========================
#             # NON-WCOM FLOW
#             # =========================
#             if lob.upper() != "WCOM":

#                 Commercial_AP = page.locator("text='Commercial AP'")

#                 await Commercial_AP.wait_for(state="visible", timeout=15000)

#                 await Commercial_AP.click()

#                 await page.locator("text='Commercial AP'").click()

#                 await page.wait_for_timeout(3000)

#                 # Click Status
#                 await page.wait_for_selector("//span[text()='Status']")

#                 await page.locator("//span[text()='Status']").click()

#                 # =========================
#                 # GET PREMIUM
#                 # =========================

#                 locator = page.locator('#curePolicyPremium__textField')

#                 await locator.wait_for(state="attached", timeout=15000)

#                 # Scroll until visible
#                 for _ in range(5):

#                     await locator.scroll_into_view_if_needed()

#                     if await locator.is_visible():
#                         break

#                     await page.mouse.wheel(0, 1500)

#                     await page.wait_for_timeout(1000)

#                 await locator.wait_for(state="visible", timeout=20000)

#                 premium = ""

#                 # Wait until premium value loads
#                 for _ in range(15):

#                     try:
#                         premium = (await locator.input_value()).strip()
#                     except:
#                         premium = (await locator.text_content() or "").strip()

#                     if premium:
#                         break

#                     print("Waiting for premium value...")

#                     await page.wait_for_timeout(1000)

#                 if premium and premium != "$":

#                     if not premium.startswith("$"):
#                         premium = f"${premium}"

#                     print("Premium Value:", premium)

#                     return premium

#                 return "Premium not found"

#             # =========================
#             # WCOM FLOW
#             # =========================
#             else:

#                 premium = ""

#                 sections = [
#                     {
#                         "name": "Policy Information/Total Premiums",
#                         "locator": '#cureTotalEstimatedAnnualPremium__textField'
#                     },
#                     {
#                         "name": "Total Premium Calculations",
#                         "locator": '[data-automation-id="cureTotalAnnualPremium"] input'
#                     }
#                 ]

#                 for section in sections:

#                     section_name = section["name"]

#                     locator_value = section["locator"]
#                     try:
#                         print(f"Opening section: {section_name}")

#                         section_tab = page.locator(f"text='{section_name}'")

#                         await section_tab.wait_for(state="visible", timeout=15000)

#                         await section_tab.click()

#                         locator = page.locator(locator_value)

#                         # Wait until attached
#                         await locator.wait_for(state="attached", timeout=15000)

#                         # Scroll until visible
#                         for _ in range(5):

#                             await locator.scroll_into_view_if_needed()

#                             if await locator.is_visible():
#                                 break

#                             await page.mouse.wheel(0, 1500)

#                             await page.wait_for_timeout(1000)

#                         await locator.wait_for(state="visible", timeout=5000)

#                         # Wait until premium value loads
#                         for _ in range(15):

#                             try:
#                                 premium = (await locator.input_value()).strip()
#                             except:
#                                 premium = (await locator.text_content() or "").strip()

#                             if premium:
#                                 break

#                             print("Waiting for premium value...")

#                             await page.wait_for_timeout(1000)

#                         if premium and premium != "$":

#                             if not premium.startswith("$"):
#                                 premium = f"${premium}"

#                             print("Premium Value:", premium)

#                             return premium

#                     except Exception as e:

#                         print(f"Premium not found in section {section}: {e}")

#                         continue

#                 return "Premium not found"

# #SCRAPER FUNCTION

# async def scrape_records(page, start_id):

#     locator = page.get_by_text("Follow Up/Start")
#     await locator.wait_for(state="visible", timeout=20000)

#     await page.get_by_text("Customize View").click()
#     await page.locator('[data-automation-id="tpgActivitiesDisplayTab"]').click()
#     await page.get_by_text("Selected date range").click()

#     yesterday = datetime.now() - timedelta(days=2)
#     yesterday_str = f"{yesterday.month}/{yesterday.day}/{yesterday.year}"

#     print("Yesterday:", yesterday_str)

#     date_inputs = page.locator("input.date-edit[type='text']")
#     await date_inputs.first.wait_for(state="visible")

#     await date_inputs.nth(0).fill(yesterday_str)
#     await date_inputs.nth(1).fill(yesterday_str)

#     await page.get_by_role("button", name="Finish").click()
#     await page.wait_for_timeout(5000)

#     data_list = []
#     row = 1
#     current_id = start_id

#     MAX_RETRIES = 3

#     while True:

#         await page.wait_for_timeout(5000)

#         print(f"Processing row {row}")

#         data_visible = f"/html/body/root/app/program/div/div[2]/div[2]/div/div/screen/div/div/div/proxy[3]/div/asi-panel/div/proxy[3]/div/asi-frame/div/div[2]/proxy/div/asi-frame/div/div[2]/proxy[3]/div/asi-virtual-list-view/div/div/asi-virtual-list-table/div/div/div[2]/div[2]/div/div[{row}]/div[5]/div/span"
#         locator = page.locator(f"xpath={data_visible}")

#         try:
#             await locator.wait_for(state="visible", timeout=5000)
#         except:
#             print(f"Row {row} not visible — checking scroll...")

#             await page.mouse.wheel(0, 2000)
#             await page.wait_for_timeout(2000)

#             if await locator.count() == 0:
#                 print("Confirmed end of rows")
#                 return data_list
#             else:
#                 print("Row loaded after scroll, continuing...")

#         await locator.wait_for(state="visible", timeout=20000)

#         date_xpath = f"/html/body/root/app/program/div/div[2]/div[2]/div/div/screen/div/div/div/proxy[3]/div/asi-panel/div/proxy[3]/div/asi-frame/div/div[2]/proxy/div/asi-frame/div/div[2]/proxy[3]/div/asi-virtual-list-view/div/div/asi-virtual-list-table/div/div/div[2]/div[2]/div/div[{row}]/div[5]/div/span"

#         date_element = page.locator(f"xpath={date_xpath}")
#         await date_element.wait_for(timeout=5000)

#         date_text = (await date_element.inner_text()).strip()
#         print("Date Text:", date_text)
        
#         if date_text != yesterday_str:
#             row += 1
#             continue

#         # -------- CREATE REF ID ONCE --------

#         timestamp = datetime.now().strftime("%Y%m%d")
#         ref_id = f"REF{str(current_id).zfill(6)}_{timestamp}"
#         current_id += 1
        
        
#         task_folder = os.path.join(PDF_FOLDER, ref_id)
#         os.makedirs(task_folder, exist_ok=True)

#         retry = 0

#         status = "Pending"

#         while retry < MAX_RETRIES:

#             try:

#                 await page.wait_for_timeout(5000)

#                 print("Processing Row:", row, "| Ref ID:", ref_id)

#                 task_folder = os.path.join(PDF_FOLDER, ref_id)
#                 os.makedirs(task_folder, exist_ok=True)

#                 desc_xpath = f"/html/body/root/app/program/div/div[2]/div[2]/div/div/screen/div/div/div/proxy[3]/div/asi-panel/div/proxy[3]/div/asi-frame/div/div[2]/proxy/div/asi-frame/div/div[2]/proxy[3]/div/asi-virtual-list-view/div/div/asi-virtual-list-table/div/div/div[2]/div[2]/div/div[{row}]/div[1]/div/span"

#                 desc_ele = page.locator(f"xpath={desc_xpath}")
#                 await desc_ele.wait_for(timeout=10000)
#                 await desc_ele.scroll_into_view_if_needed()

#                 row_text = await desc_ele.inner_text()
#                 print("Row clicked:", row_text)

#                 await desc_ele.click(force=True)

#                 await page.wait_for_timeout(10000)

#                 open_activity = page.get_by_text("Open Activity", exact=True)
#                 await open_activity.wait_for(state="visible", timeout=20000)

#                 description = page.locator("//asi-string-edit//div").first
#                 await description.wait_for(timeout=10000)
#                 description_value = await description.get_attribute("data-value")
#                 print("Description:", description_value)

#                 await page.wait_for_timeout(5000)

#                 association = page.locator('[data-automation-id="plhAssociation"] .text')
#                 await association.wait_for(timeout=10000)
#                 association_value = await association.text_content()
#                 association_value = association_value.strip() if association_value else ""

#                 print("Association:", association_value)

#                 parts = [p.strip() for p in association_value.split("-", 2)]

#                 lob = parts[1] if len(parts) > 1 else ""
#                 policy_no = parts[2] if len(parts) > 2 else ""

#                 print("LOB:", lob)
#                 print("Policy Number:", policy_no)

#                 await page.locator("text=Account Detail").click()


#                 account_name = page.locator("//div[text()='Account name']/following::div[@data-value][1]")
#                 account_name_value = await account_name.get_attribute("data-value")
#                 print("Account Name:", account_name_value)

#                 lookup_code = page.locator("//div[text()='Lookup code']/following::div[@data-value][1]")
#                 lookup_code_value = await lookup_code.get_attribute("data-value")
#                 print("Lookup Code:", lookup_code_value)

#                 await page.locator("[data-automation-id='tpgServicingTab']").click()

#                 account_manager = page.locator('[data-automation-id="cboEmployee1"] input')
#                 await account_manager.wait_for(timeout=10000)
#                 account_manager_value = await account_manager.input_value()
#                 print("Account Manager:", account_manager_value)
                

#                 #-------------------------------------     Premium & CSR Type Extraction Logic     ---------------------------------------------
                
#                 await page.locator("//span[normalize-space()='Policies']").click()

#                 premium = await policy_premium(page, lob, policy_no)
#                 print("Captured Premium:", premium)

#                 await page.wait_for_timeout(3000)

#                 await page.locator("//span[normalize-space()='Activities']").click(timeout=5000)
#                 print("Activities clicked")

#                 await page.wait_for_timeout(5000)

#                 access = page.locator("div.main-button", has_text="Access")
#                 await access.wait_for()
#                 await access.click()
#                 print("Access clicked to download attachments")
    
#                 await page.locator("span.text.force-underline", has_text="Attachments").first.click()
#                 print("Attachments clicked")

#                 try:
#                     await page.wait_for_selector(
#                         "//div[contains(@data-automation-id,'vlvwAttachment body-row')]",
#                         state="visible",
#                         timeout=10000
#                     )
#                 except:
#                     await page.wait_for_timeout(20000)


#                 rows = page.locator("//div[contains(@data-automation-id,'vlvwAttachment body-row')]")
#                 row_count = await rows.count()

#                 print("Total Files:", row_count)

#                 for i in range(row_count):

#                     retry = 0
#                     while retry < 3:
#                         try:

#                             file_xpath = f"(//div[contains(@data-automation-id,'vlvwAttachment body-row')]//div[contains(@class,'body-cell first')]//span[contains(@class,'text')])[{i+1}]"
#                             file_element = page.locator(f"xpath={file_xpath}")
#                             await file_element.wait_for(state="visible", timeout=15000)

#                             file_text = (await file_element.inner_text()).strip()
#                             print("Processing:", file_text)

#                             await file_element.scroll_into_view_if_needed()
#                             await file_element.dblclick()

#                             download_btn = page.locator("[data-automation-id='Download']")
#                             await download_btn.wait_for(state="visible")

#                             async with page.expect_download(timeout=20000) as d:
#                                 await download_btn.click()

#                             download = await d.value

#                             file_name = re.sub(r'[<>:"/\\|?*]', '-', download.suggested_filename)

#                             save_path = os.path.join(task_folder, file_name)
#                             await download.save_as(save_path)

#                             print("Saved:", save_path)

#                             if save_path.lower().endswith(".msg"):
#                                 extract_pdf_from_msg(save_path)

#                             await page.wait_for_timeout(3000)

#                             break  

#                         except Exception as e:
#                             retry += 1
#                             print(f"Error at row {i+1} retry {retry}: {e}")
#                             await page.wait_for_timeout(3000)


#                 data_list.append({
#                     "Date": date_text,
#                     "Account Manager": account_manager_value,
#                     "Lookup Code": lookup_code_value,
#                     "Account Name": account_name_value,
#                     "Description": description_value,
#                     "Policy number": policy_no,
#                     "LOB": lob,
#                     "Association": association_value,
#                     "Premium": premium,       
#                     "CSR Type": "", 
#                     "Reference ID": ref_id,
#                     "Job ID": "",
#                     "Status": "Pending"
#                 })

#                 # await page.locator("div.close-button").wait_for(state="visible")
#                 # await page.locator("div.close-button").click()

#                 # await page.locator("div.main-button:has-text('Home')").click()
#                 # print("Clicked Home button")
#                 close_btn = page.locator("div.close-button[title*='Close']").last
#                 await close_btn.wait_for(state="visible", timeout=10000)
#                 await close_btn.click()

#                 break

#             except Exception as e:
                    
#                     await page.wait_for_timeout(10000)

#                     retry += 1
#                     print(f"\n Error at row {row} retry {retry}")
#                     traceback.print_exc()

#                     home_btn = page.locator("div.main-button:has-text('Home')")
   
#                     if await home_btn.is_visible():
#                         await home_btn.click()
#                         print("Clicked Home before retry")

#                         await page.get_by_text("Follow Up/Start").wait_for(state="visible", timeout=20000)

#                     await page.wait_for_timeout(3000)

#                     if retry >= MAX_RETRIES:
#                         print(f"Skipping row {row} after {MAX_RETRIES} retries")

#                         data_list.append({
#                             "Date": date_text,
#                             "Account Manager": "",
#                             "Lookup Code": "",
#                             "Account Name": "",
#                             "Description": "",
#                             "Policy number": "",
#                             "LOB": "",
#                             "Association": "",
#                             "Premium": "",       
#                             "CSR Type": "", 
#                             "Reference ID": ref_id,
#                             "Job ID": "",
#                             "Status": "Error"

#                         })

#                         break
                
#         row += 1

#     return data_list



# # SAVE TO EXCEL FUNCTION

# def save_to_excel(data_list):

#     new_df = pd.DataFrame(data_list)


#     if os.path.exists(EXCEL_FILE_PATH):
#         old_df = pd.read_excel(EXCEL_FILE_PATH)
#         final_df = pd.concat([old_df, new_df], ignore_index=True)
#     else:
#         final_df = new_df
    
#     final_df["Policy number"] = final_df["Policy number"].astype(str)
    
#     with pd.ExcelWriter(EXCEL_FILE_PATH, engine="openpyxl") as writer:
#         final_df.to_excel(writer, index=False, sheet_name="Sheet1")

#         worksheet = writer.sheets["Sheet1"]

#         for i, col in enumerate(final_df.columns, 1):
#             max_len = max(final_df[col].astype(str).map(len).max(), len(col)) + 2
#             worksheet.column_dimensions[get_column_letter(i)].width = max_len
    
#     print("Excel Saved Successfully")


# # MAIN FUNCTION


# async def run_epic():

#     start_id = get_starting_ref_id()

#     browser = None
#     context = None
#     page = None

#     async with async_playwright() as p:

#         try:
#             browser = await p.chromium.launch(channel="chrome", headless=False)
#             context = await browser.new_context()
#             page = await context.new_page()

#             page.set_default_timeout(30000)

#             await login(page, context)
#             data_list = await scrape_records(page, start_id)
#             save_to_excel(data_list)

#         finally:
#             print("Cleaning EPIC resources...")

#             if page:
#                 try:
#                     await page.get_by_text("Logout").click(timeout=5000)
#                     await page.locator("[data-automation-id='Yes']").click(timeout=5000)
#                     await page.wait_for_timeout(3000)
#                     print("Logout successful")
#                 except Exception as e:
#                     print("Logout skipped:", e)

#             if context:
#                 await context.close()

#             if browser:
#                 await browser.close()

# if __name__ == "__main__":
#     asyncio.run(run_epic())





#------------------------------UPDATED CODE---------------------------------#


# import asyncio
# from email import policy
# from playwright.async_api import async_playwright
# from configparser import ConfigParser
# from datetime import datetime, timedelta
# import pandas as pd
# import os
# import re
# import extract_msg
# from openpyxl.utils import get_column_letter
# import traceback
# from playwright.async_api import Error as PlaywrightError

# config = ConfigParser()
# config.read("c:/Users/vijay_m/OneDrive - Exdion Solutions Pvt. Ltd.-70692290/Documents/Project_Job_Creation/Thomson/config.ini")
# BASE_FOLDER = config.get("PATHS", "BASE_FOLDER")
# PDF_FOLDER = os.path.join(BASE_FOLDER, "Job_Creation")
# EXCEL_FILE_PATH = os.path.join(BASE_FOLDER, "Thomson_Tracker.xlsx")
# CHECKPOINT_FILE = os.path.join(BASE_FOLDER, "epic_checkpoint.txt")

# os.makedirs(BASE_FOLDER, exist_ok=True)
# os.makedirs(PDF_FOLDER, exist_ok=True)

# # ================= GET STARTING REFERENCE NUMBER =================

# def get_starting_ref_id():

#     if not os.path.exists(EXCEL_FILE_PATH):
#         return 1

#     df = pd.read_excel(EXCEL_FILE_PATH)

#     if "Reference ID" not in df.columns or df.empty:
#         return 1

#     existing_ids = df["Reference ID"].dropna()

#     numbers = (
#         existing_ids
#         .str.extract(r'REF(\d+)')[0]
#         .dropna()
#         .astype(int)
#     )

#     return numbers.max() + 1 if not numbers.empty else 1



# # LOGIN FUNCTION

# async def login(page, context):

#     url = config.get("ThomsonEpic", "Link")
#     EnterpriseID = config.get("ThomsonEpic", "Enterprise_ID")
#     username = config.get("ThomsonEpic", "User_ID")
#     password = config.get("ThomsonEpic", "Password")

#     await page.goto(url, timeout=90000)

#     await page.get_by_role("textbox").fill(EnterpriseID)

#     await page.locator("//button[.//span[contains(text(),'Contin')]]").click()

#     async with context.expect_page() as new_page_event:
#         await page.locator("//button[contains(text(),'Login')]").click()

#     login_page = await new_page_event.value

#     await login_page.locator("#usercode").fill(username)
#     await login_page.locator("#password").fill(password)

#     await login_page.locator("//button[.//div[text()='Login']]").click()

#     try:
#         later_btn = login_page.get_by_role("button", name="Later")
#         #login_page.locator("button:has-text('Later')").click()
#         await later_btn.wait_for(timeout=5000)
#         await later_btn.click()
#         print("Later Reset clicked")
#     except:
#         pass
    
#     await page.wait_for_load_state("networkidle")

#     await page.locator("//div[contains(@class,'drop-btn')]").first.click()

#     await page.locator("//span[@class='text' and text()='TSLIN01_PROD']").click()

#     await page.locator("//button[.//span[contains(text(),'Contin')]]").click()

#     try:
#         await page.locator("//button[@data-automation-id='Yes']").click(timeout=3000)
#         print("Popup YES clicked, Login successful")
#     except:
#         print("Login successful")

#     # await page.pause()

#     return login_page


# #--------------------------------------------It Works if .msg file in last-----------------------------------------------------------------
# import os
# import extract_msg

# def extract_pdf_from_msg(msg_path, existing_types):
#     try:
#         msg = extract_msg.Message(msg_path)
#         save_folder = os.path.dirname(msg_path)

#         target_keywords = [
#             "quote", "binder", "proposal",
#             "schedule", "application", "app", "accord"
#         ]

#         for att in msg.attachments:
#             filename = att.longFilename or att.shortFilename
#             if not filename:
#                 continue

#             filename_lower = filename.lower()

#             if filename_lower.endswith(".pdf"):

#                 matched_keyword = next((k for k in target_keywords if k in filename_lower), None)

#                 if matched_keyword:
#                     if matched_keyword in existing_types:
#                         print(f"Skipping INNER {matched_keyword}:", filename)
#                         continue

#                     pdf_path = os.path.join(save_folder, filename)

#                     with open(pdf_path, "wb") as f:
#                         f.write(att.data)

#                     print("Saved INNER:", pdf_path)

#                     existing_types.add(matched_keyword)

#     except Exception as e:
#         print("MSG Extraction Error:", e)

# #--------------------Checkpoint functions to save progress and avoid duplicates in case of crash-------------------#

# def save_checkpoint(row):
#     with open(CHECKPOINT_FILE, "w") as f:
#         f.write(str(row))

# def load_checkpoint():
#     if not os.path.exists(CHECKPOINT_FILE):
#         return 1
#     try:
#         with open(CHECKPOINT_FILE, "r") as f:
#             return int(f.read().strip())
#     except:
#         return 1

# def clear_checkpoint():
#     if os.path.exists(CHECKPOINT_FILE):
#         os.remove(CHECKPOINT_FILE)        

# #-------------------------------------     Premium Scrap Extraction Logic     ---------------------------------------------
        
# from datetime import datetime
# from playwright.async_api import expect

# async def policy_premium(page, lob, policy_no, exp_eff_year=2026, exp_exp_year=2027):

#     rows = page.locator('[data-automation-id*="vlvwPolicies body-row"]')

#     await rows.first.wait_for(timeout=15000)

#     count = await rows.count()
#     print("Row count:", count)

#     for i in range(count):

#         row = rows.nth(i)
#         cells = row.locator('.body-cell .text')

#         try:
#             line = (await cells.nth(0).text_content() or "").strip()
#             effective = (await cells.nth(2).text_content() or "").strip()
#             expiration = (await cells.nth(3).text_content() or "").strip()
#             policy_number = (await cells.nth(4).text_content() or "").strip()

#             #print(f"Checking Row {i}: LOB={line}, Effective={effective}, Expiration={expiration}")

#             print(
#                 f"Checking Row {i}: "
#                 f"LOB={line}, "
#                 f"Policy={policy_number}, "
#                 f"Effective={effective}, "
#                 f"Expiration={expiration}"
#             )

#             if lob.lower() not in line.lower():
#                 continue

#             if policy_no.strip().lower() != policy_number.lower():
#                 continue

#             try:
#                 eff_year = datetime.strptime(effective, "%m/%d/%Y").year
#                 exp_year = datetime.strptime(expiration, "%m/%d/%Y").year
#             except:
#                 continue

#         except Exception as e:
#             print("Row Error:", e)
#             continue

#         if eff_year == exp_eff_year and exp_year == exp_exp_year:

#             print(f"Match found at row {i}, opening...")

#             await row.dblclick()

#             # =========================
#             # NON-WCOM FLOW
#             # =========================
#             if lob.upper() != "WCOM":

#                 Commercial_AP = page.locator("text='Commercial AP'")

#                 await Commercial_AP.wait_for(state="visible", timeout=15000)

#                 await Commercial_AP.click()

#                 #await page.locator("text='Commercial AP'").click()

#                 await page.wait_for_timeout(3000)

#                 # Click Status
#                 await page.wait_for_selector("//span[text()='Status']")

#                 await page.locator("//span[text()='Status']").click()

#                 # =========================
#                 # GET PREMIUM
#                 # =========================

#                 locator = page.locator('#curePolicyPremium__textField')

#                 await locator.wait_for(state="attached", timeout=15000)

#                 # Scroll until visible
#                 for _ in range(5):

#                     await locator.scroll_into_view_if_needed()

#                     if await locator.is_visible():
#                         break

#                     await page.mouse.wheel(0, 1500)

#                     await page.wait_for_timeout(1000)

#                 await locator.wait_for(state="visible", timeout=20000)

#                 premium = ""

#                 # Wait until premium value loads
#                 for _ in range(25):

#                     try:
#                         premium = (await locator.input_value()).strip()
#                     except:
#                         premium = (await locator.text_content() or "").strip()

#                     if premium:
#                         break

#                     print("Waiting for premium value...")

#                     await page.wait_for_timeout(1000)

#                 if premium and premium != "$":

#                     if not premium.startswith("$"):
#                         premium = f"${premium}"

#                     print("Premium Value:", premium)

#                     return premium

#                 return "Premium not found"

#             # =========================
#             # WCOM FLOW
#             # =========================
#             else:

#                 premium = ""

#                 sections = [
#                     {
#                         "name": "Policy Information/Total Premiums",
#                         "locator": '#cureTotalEstimatedAnnualPremium__textField'
#                     },
#                     {
#                         "name": "Total Premium Calculations",
#                         "locator": '[data-automation-id="cureTotalAnnualPremium"] input'
#                     }
#                 ]

#                 for section in sections:

#                     section_name = section["name"]

#                     locator_value = section["locator"]
#                     try:
#                         print(f"Opening section: {section_name}")

#                         section_tab = page.locator(f"text='{section_name}'")

#                         await section_tab.wait_for(state="visible", timeout=15000)

#                         await section_tab.click()

#                         locator = page.locator(locator_value)

#                         # Wait until attached
#                         await locator.wait_for(state="attached", timeout=15000)

#                         # Scroll until visible
#                         for _ in range(5):

#                             await locator.scroll_into_view_if_needed()

#                             if await locator.is_visible():
#                                 break

#                             await page.mouse.wheel(0, 1500)

#                             await page.wait_for_timeout(1000)

#                         await locator.wait_for(state="visible", timeout=5000)

#                         # Wait until premium value loads
#                         for _ in range(15):

#                             try:
#                                 premium = (await locator.input_value()).strip()
#                             except:
#                                 premium = (await locator.text_content() or "").strip()

#                             if premium:
#                                 break

#                             print("Waiting for premium value...")

#                             await page.wait_for_timeout(1000)

#                         if premium and premium != "$":

#                             if not premium.startswith("$"):
#                                 premium = f"${premium}"

#                             print("Premium Value:", premium)

#                             return premium

#                     except Exception as e:

#                         print(f"Premium not found in section {section}: {e}")

#                         continue

#                 return "Premium not found"

# #SCRAPER FUNCTION

# async def scrape_records(page, start_id, existing_ids, target_date):

#     locator = page.get_by_text("Follow Up/Start")
#     await locator.wait_for(state="visible", timeout=20000)

#     await page.get_by_text("Customize View").click()
#     await page.locator('[data-automation-id="tpgActivitiesDisplayTab"]').click()
#     await page.get_by_text("Selected date range").click()

#     # yesterday = datetime.now() - timedelta(days=1)
#     # yesterday_str = f"{yesterday.month}/{yesterday.day}/{yesterday.year}"
#     yesterday_str = f"{target_date.month}/{target_date.day}/{target_date.year}"

#     print("Yesterday:", yesterday_str)

#     date_inputs = page.locator("input.date-edit[type='text']")
#     await date_inputs.first.wait_for(state="visible")

#     await date_inputs.nth(0).fill(yesterday_str)
#     await date_inputs.nth(1).fill(yesterday_str)

#     await page.get_by_role("button", name="Finish").click()

#     await page.wait_for_timeout(5000)

#     locator = page.get_by_text("Follow Up/Start")
#     await locator.wait_for(state="visible", timeout=20000)

#     row = load_checkpoint()
#     print(f"▶ Resuming from row: {row}")

#     current_id = start_id

#     MAX_RETRIES = 3

#     while True:

#         if page.is_closed():
#             raise Exception("Page closed → restart browser")

#         data_visible = f"/html/body/root/app/program/div/div[2]/div[2]/div/div/screen/div/div/div/proxy[3]/div/asi-panel/div/proxy[3]/div/asi-frame/div/div[2]/proxy/div/asi-frame/div/div[2]/proxy[3]/div/asi-virtual-list-view/div/div/asi-virtual-list-table/div/div/div[2]/div[2]/div/div[{row}]/div[5]/div/span"
#         locator = page.locator(f"xpath={data_visible}")

#         #await page.wait_for_timeout(5000)

#         try:
#             await locator.wait_for(state="visible", timeout=10000)
#         except:
#             print(f"Row {row} not visible — checking scroll...")

#             await page.mouse.wheel(0, 2000)
#             await page.wait_for_timeout(2000)

#             if await locator.count() == 0:
#                 print("Confirmed end of rows")
#                 return
#             else:
#                 print("Row loaded after scroll, continuing...")

#         await locator.wait_for(state="visible", timeout=20000)

#         date_xpath = f"/html/body/root/app/program/div/div[2]/div[2]/div/div/screen/div/div/div/proxy[3]/div/asi-panel/div/proxy[3]/div/asi-frame/div/div[2]/proxy/div/asi-frame/div/div[2]/proxy[3]/div/asi-virtual-list-view/div/div/asi-virtual-list-table/div/div/div[2]/div[2]/div/div[{row}]/div[5]/div/span"

#         date_element = page.locator(f"xpath={date_xpath}")
#         await date_element.wait_for(timeout=5000)

#         date_text = (await date_element.inner_text()).strip()
#         print("Date Text:", date_text)
        
#         if date_text != yesterday_str:
#             row += 1
#             continue

#         # -------- CREATE REF ID ONCE --------

#         timestamp = datetime.now().strftime("%Y%m%d")
#         ref_id = f"REF{str(current_id).zfill(6)}_{timestamp}"
#         current_id += 1
        
        
#         task_folder = os.path.join(PDF_FOLDER, ref_id)
#         os.makedirs(task_folder, exist_ok=True)

#         retry = 0

#         status = "Pending"

#         while retry < MAX_RETRIES:

#             if page.is_closed():
#                 raise Exception("Page closed → restart browser")

#             try:

#                 await page.wait_for_timeout(5000)

#                 print("Processing Row:", row, "| Ref ID:", ref_id)

#                 task_folder = os.path.join(PDF_FOLDER, ref_id)
#                 os.makedirs(task_folder, exist_ok=True)

#                 desc_xpath = f"/html/body/root/app/program/div/div[2]/div[2]/div/div/screen/div/div/div/proxy[3]/div/asi-panel/div/proxy[3]/div/asi-frame/div/div[2]/proxy/div/asi-frame/div/div[2]/proxy[3]/div/asi-virtual-list-view/div/div/asi-virtual-list-table/div/div/div[2]/div[2]/div/div[{row}]/div[1]/div/span"

#                 desc_ele = page.locator(f"xpath={desc_xpath}")
#                 await desc_ele.wait_for(timeout=10000)
#                 await desc_ele.scroll_into_view_if_needed()

#                 row_text = await desc_ele.inner_text()
#                 print("Row clicked:", row_text)

#                 await desc_ele.click(force=True)

#                 await page.wait_for_timeout(10000)

#                 open_activity = page.get_by_text("Open Activity", exact=True)
#                 await open_activity.wait_for(state="visible", timeout=20000)

#                 description = page.locator("//asi-string-edit//div").first
#                 await description.wait_for(timeout=10000)
#                 description_value = await description.get_attribute("data-value")
#                 print("Description:", description_value)

#                 await page.wait_for_timeout(5000)

#                 association = page.locator('[data-automation-id="plhAssociation"] .text')
#                 await association.wait_for(timeout=10000)
#                 association_value = await association.text_content()
#                 association_value = association_value.strip() if association_value else ""

#                 print("Association:", association_value)

#                 parts = [p.strip() for p in association_value.split(" - ", 2)]

#                 lob = parts[1] if len(parts) > 1 else ""
#                 policy_no = parts[2] if len(parts) > 2 else ""

#                 print("LOB:", lob)
#                 print("Policy Number:", policy_no)

#                 # ===== CREATE UNIQUE ID =====

#                 unique_id = association_value.strip().lower()

#                 print("CHECK UNIQUE:", unique_id)

#                 if unique_id in existing_ids:

#                     print("Skipping Duplicate:", unique_id)

#                     close_btn = page.locator("div.close-button[title*='Close']").last

#                     await close_btn.wait_for(state="visible", timeout=10000)

#                     await close_btn.click()

#                     # await page.locator(
#                     #     "div.main-button:has-text('Home')"
#                     # ).click()

#                     row += 1
#                     continue


#                 await page.locator("text=Account Detail").click()


#                 account_name = page.locator("//div[text()='Account name']/following::div[@data-value][1]")
#                 account_name_value = await account_name.get_attribute("data-value")
#                 print("Account Name:", account_name_value)

#                 lookup_code = page.locator("//div[text()='Lookup code']/following::div[@data-value][1]")
#                 lookup_code_value = await lookup_code.get_attribute("data-value")
#                 print("Lookup Code:", lookup_code_value)

#                 await page.locator("[data-automation-id='tpgServicingTab']").click()

#                 account_manager = page.locator('[data-automation-id="cboEmployee1"] input')
#                 await account_manager.wait_for(timeout=10000)
#                 account_manager_value = await account_manager.input_value()
#                 print("Account Manager:", account_manager_value)
                

#                 #-------------------------------------     Premium & CSR Type Extraction Logic     ---------------------------------------------
                
#                 await page.locator("//span[normalize-space()='Policies']").click()

#                 premium = await policy_premium(page, lob, policy_no)
#                 print("Captured Premium:", premium)

#                 await page.wait_for_timeout(3000)

#                 #--------------------------------------------------------------------------------------------------------------------------------

#                 await page.locator("//span[normalize-space()='Activities']").click(timeout=5000)
#                 print("Activities clicked")

#                 await page.wait_for_timeout(5000)

#                 access = page.locator("div.main-button", has_text="Access")
#                 await access.wait_for()
#                 await access.click()
#                 print("Access clicked to download attachments")
    
#                 await page.locator("span.text.force-underline", has_text="Attachments").first.click()
#                 print("Attachments clicked")

#                 try:
#                     await page.wait_for_selector(
#                         "//div[contains(@data-automation-id,'vlvwAttachment body-row')]",
#                         state="visible",
#                         timeout=10000
#                     )
#                 except:
#                     await page.wait_for_timeout(20000)


#                 rows = page.locator("//div[contains(@data-automation-id,'vlvwAttachment body-row')]")
#                 row_count = await rows.count()

#                 print("Total Files:", row_count)

#                 for i in range(row_count):

#                     retry = 0
#                     while retry < 3:
#                         try:

#                             file_xpath = f"(//div[contains(@data-automation-id,'vlvwAttachment body-row')]//div[contains(@class,'body-cell first')]//span[contains(@class,'text')])[{i+1}]"
#                             file_element = page.locator(f"xpath={file_xpath}")
#                             await file_element.wait_for(state="visible", timeout=15000)

#                             file_text = (await file_element.inner_text()).strip()
#                             print("Processing:", file_text)

#                             await file_element.scroll_into_view_if_needed()
#                             await file_element.dblclick()

#                             download_btn = page.locator("[data-automation-id='Download']")
#                             await download_btn.wait_for(state="visible")

#                             async with page.expect_download(timeout=20000) as d:
#                                 await download_btn.click()

#                             download = await d.value

#                             file_name = re.sub(r'[<>:"/\\|?*]', '-', download.suggested_filename)

#                             save_path = os.path.join(task_folder, file_name)
#                             await download.save_as(save_path)

#                             print("Saved:", save_path)

#                             if save_path.lower().endswith(".msg"):
#                                 extract_pdf_from_msg(save_path)

#                             await page.wait_for_timeout(3000)

#                             break  

#                         except Exception as e:
#                             retry += 1
#                             print(f"Error at row {i+1} retry {retry}: {e}")
#                             await page.wait_for_timeout(3000)
                

#                 # =========================
#                 # FINAL STEP: Process MSG after ALL downloads
#                 # =========================

#                 target_keywords = [
#                     "quote", "binder", "proposal",
#                     "schedule", "application", "app", "accord"
#                 ]

#                 existing_types = set()

#                 # Detect OUTER PDFs
#                 for f in os.listdir(task_folder):
#                     f_lower = f.lower()
#                     if f_lower.endswith(".pdf"):
#                         for k in target_keywords:
#                             if k in f_lower:
#                                 existing_types.add(k)

                
#                 msg_files = [f for f in os.listdir(task_folder) if f.lower().endswith(".msg")]

#                 if msg_files:
#                     print("Existing OUTER types:", existing_types)

#                     for f in msg_files:
#                         msg_path = os.path.join(task_folder, f)
#                         extract_pdf_from_msg(msg_path, existing_types)

#                 row_data = {
#                     "Date": date_text,
#                     "Account Manager": account_manager_value,
#                     "Lookup Code": lookup_code_value,
#                     "Account Name": account_name_value,
#                     "Description": description_value,
#                     "Policy number": policy_no,
#                     "LOB": lob,
#                     "Association": association_value,
#                     "Premium": premium,       
#                     "CSR Type": "", 
#                     "Reference ID": ref_id,
#                     "Job ID": "",
#                     "Status": "Pending"
#                 }

#                 append_to_excel(row_data, existing_ids)
#                 existing_ids.add(unique_id)
#                 save_checkpoint(row)

#                 # await page.locator("div.close-button").wait_for(state="visible")
#                 # await page.locator("div.close-button").click()

#                 # await page.locator("div.main-button:has-text('Home')").click()
#                 # print("Clicked Home button")
#                 close_btn = page.locator("div.close-button[title*='Close']").last
#                 await close_btn.wait_for(state="visible", timeout=10000)
#                 await close_btn.click()

#                 break

#             except Exception as e:
                    
#                     await page.wait_for_timeout(10000)

#                     retry += 1
#                     print(f"\n Error at row {row} retry {retry}")
#                     traceback.print_exc()

#                     home_btn = page.locator("div.main-button:has-text('Home')")
   
#                     if await home_btn.is_visible():
#                         await home_btn.click()
#                         print("Clicked Home before retry")

#                         await page.get_by_text("Follow Up/Start").wait_for(state="visible", timeout=20000)

#                     await page.wait_for_timeout(3000)

#                     if retry >= MAX_RETRIES:
#                         print(f"Skipping row {row} after {MAX_RETRIES} retries")

#                         error_data = {
#                             "Date": date_text,
#                             "Account Manager": account_manager_value,
#                             "Lookup Code": lookup_code_value,
#                             "Account Name": account_name_value,
#                             "Description": description_value,
#                             "Policy number": policy_no,
#                             "LOB": lob,
#                             "Association": association_value,
#                             "Premium": "",       
#                             "CSR Type": "", 
#                             "Reference ID": ref_id,
#                             "Job ID": "",
#                             "Status": "Error"

#                         }

#                         append_to_excel(error_data, existing_ids)
#                         save_checkpoint(row)
#                         break
                
#         row += 1

# # SAVE TO EXCEL FUNCTION

# def append_to_excel(row_data, existing_ids):

#     row_data["Reference ID"] = str(row_data["Reference ID"])

#     df = pd.DataFrame([row_data])

#     if not os.path.exists(EXCEL_FILE_PATH):
#         df.to_excel(EXCEL_FILE_PATH, index=False)
#         print(f"Saved → {row_data['Reference ID']}")
#         return True

#     from openpyxl import load_workbook
#     book = load_workbook(EXCEL_FILE_PATH)
#     sheet = book.active
#     startrow = sheet.max_row
#     book.close()

#     with pd.ExcelWriter(
#         EXCEL_FILE_PATH,
#         engine="openpyxl",
#         mode="a",
#         if_sheet_exists="overlay"
#     ) as writer:

#         df.to_excel(
#             writer,
#             index=False,
#             header=False,
#             startrow=startrow
#         )

#     print(f"Saved → {row_data['Reference ID']}")
#     return True



# # MAIN FUNCTION

# async def run_epic():

#     MAX_BROWSER_RETRIES = 3

#     start_id = get_starting_ref_id()

#     for attempt in range(MAX_BROWSER_RETRIES):

#         browser = None
#         context = None
#         page = None

#         try:
#             print(f"\n--------------- EPIC Attempt {attempt+1}")

#             existing_ids = set()

#             if os.path.exists(EXCEL_FILE_PATH):
#                 df = pd.read_excel(EXCEL_FILE_PATH)

#                 df["Date"] = pd.to_datetime(df["Date"], errors="coerce").dt.date
#                 target_date = (datetime.now() - timedelta(days=0)).date()

#                 df_filtered = df[df["Date"] == target_date].copy()

#                 existing_ids = set(
#                     df_filtered["Association"]
#                     .astype(str)
#                     .str.strip()
#                     .str.lower()
#                 )
#             else:
#                 target_date = (datetime.now() - timedelta(days=0)).date()


#             async with async_playwright() as p:

#                     browser = await p.chromium.launch(channel="chrome", headless=False)
#                     context = await browser.new_context()
#                     page = await context.new_page()

#                     page.set_default_timeout(30000)

#                     await login(page, context)
#                     await scrape_records(page, start_id, existing_ids, target_date)

#                     if not page.is_closed():
#                         try:
#                             await page.get_by_text("Logout").click(timeout=5000)
#                             await page.locator("[data-automation-id='Yes']").click(timeout=5000)
#                             await page.wait_for_timeout(3000)
#                             print("Logout successful")
#                         except Exception as e:
#                             print("Logout failed:", e)

#                     print("EPIC completed successfully")
#                     clear_checkpoint()
#                     return 

#         # except Exception as e:
#         #     print(f"\n Browser failed: {e}")
#         #     await asyncio.sleep(5)

#         except PlaywrightError as e:
#             error_text = str(e)

#             print(f"\n Playwright error: {error_text}")

#             if "Target closed" in error_text:
#                 print("Detected browser crash → restarting...")

#             await asyncio.sleep(5)

            
#         # finally:
#         #     print("Cleaning EPIC resources...")

#         #     if page:
#         #         try:
#         #             await page.title()
#         #         except:
#         #             print("Page already closed, skipping logout")

#         #     try:
#         #         if context:
#         #             await context.close()
#         #         if browser:
#         #             await browser.close()
#         #     except:
#         #         pass

#         finally:
#             print("Cleaning EPIC resources...")

#             if page:
#                 if page.is_closed():
#                     print("Page already closed")
#                 else:
#                     print("Page still open")

#             try:
#                 if context:
#                     await context.close()
#                 if browser:
#                     await browser.close()
#             except Exception as e:
#                 print("Cleanup error:", e)

#     print("EPIC failed after max retries")

# if __name__ == "__main__":
#     asyncio.run(run_epic())




##----------------------------------------------------------------------------------------------------------------------------



import asyncio
from email import policy
from playwright.async_api import async_playwright
from configparser import ConfigParser
from datetime import datetime, timedelta
import pandas as pd
import os
import re
import extract_msg
from openpyxl.utils import get_column_letter
import traceback
from playwright.async_api import Error as PlaywrightError

config = ConfigParser()
config.read("c:/Users/vijay_m/OneDrive - Exdion Solutions Pvt. Ltd.-70692290/Documents/Project_Job_Creation/Thomson/config.ini")
BASE_FOLDER = config.get("PATHS", "BASE_FOLDER")
PDF_FOLDER = os.path.join(BASE_FOLDER, "Job_Creation")
EXCEL_FILE_PATH = os.path.join(BASE_FOLDER, "Thomson_Tracker.xlsx")
CHECKPOINT_FILE = os.path.join(BASE_FOLDER, "epic_checkpoint.txt")

os.makedirs(BASE_FOLDER, exist_ok=True)
os.makedirs(PDF_FOLDER, exist_ok=True)

# ================= GET STARTING REFERENCE NUMBER =================

def get_starting_ref_id():

    if not os.path.exists(EXCEL_FILE_PATH):
        return 1

    df = pd.read_excel(EXCEL_FILE_PATH)

    if "Reference ID" not in df.columns or df.empty:
        return 1

    existing_ids = df["Reference ID"].dropna()

    numbers = (
        existing_ids
        .str.extract(r'REF(\d+)')[0]
        .dropna()
        .astype(int)
    )

    return numbers.max() + 1 if not numbers.empty else 1



# LOGIN FUNCTION

async def login(page, context):

    url = config.get("ThomsonEpic", "Link")
    EnterpriseID = config.get("ThomsonEpic", "Enterprise_ID")
    username = config.get("ThomsonEpic", "User_ID")
    password = config.get("ThomsonEpic", "Password")

    await page.goto(url, timeout=90000)

    await page.get_by_role("textbox").fill(EnterpriseID)

    await page.locator("//button[.//span[contains(text(),'Contin')]]").click()

    async with context.expect_page() as new_page_event:
        await page.locator("//button[contains(text(),'Login')]").click()

    login_page = await new_page_event.value

    await login_page.locator("#usercode").fill(username)
    await login_page.locator("#password").fill(password)

    await login_page.locator("//button[.//div[text()='Login']]").click()

    try:
        later_btn = login_page.get_by_role("button", name="Later")
        #login_page.locator("button:has-text('Later')").click()
        await later_btn.wait_for(timeout=5000)
        await later_btn.click()
        print("Later Reset clicked")
    except:
        pass
    
    await page.wait_for_load_state("networkidle")

    await page.locator("//div[contains(@class,'drop-btn')]").first.click()

    await page.locator("//span[@class='text' and text()='TSLIN01_PROD']").click()

    await page.locator("//button[.//span[contains(text(),'Contin')]]").click()

    try:
        await page.locator("//button[@data-automation-id='Yes']").click(timeout=3000)
        print("Popup YES clicked, Login successful")
    except:
        print("Login successful")

    # await page.pause()

    return login_page


#--------------------------------------------It Works if .msg file in last-----------------------------------------------------------------
import os
import extract_msg

def extract_pdf_from_msg(msg_path, existing_types):
    try:
        msg = extract_msg.Message(msg_path)
        save_folder = os.path.dirname(msg_path)

        target_keywords = [
            "quote", "binder", "proposal",
            "schedule", "application", "app", "accord"
        ]

        for att in msg.attachments:
            filename = att.longFilename or att.shortFilename
            if not filename:
                continue

            filename_lower = filename.lower()

            if filename_lower.endswith(".pdf"):

                matched_keyword = next((k for k in target_keywords if k in filename_lower), None)

                if matched_keyword:
                    if matched_keyword in existing_types:
                        print(f"Skipping INNER {matched_keyword}:", filename)
                        continue

                    pdf_path = os.path.join(save_folder, filename)

                    with open(pdf_path, "wb") as f:
                        f.write(att.data)

                    print("Saved INNER:", pdf_path)

                    existing_types.add(matched_keyword)

    except Exception as e:
        print("MSG Extraction Error:", e)

#--------------------Checkpoint functions to save progress and avoid duplicates in case of crash-------------------#

def save_checkpoint(row):
    with open(CHECKPOINT_FILE, "w") as f:
        f.write(str(row))

def load_checkpoint():
    if not os.path.exists(CHECKPOINT_FILE):
        return 1
    try:
        with open(CHECKPOINT_FILE, "r") as f:
            return int(f.read().strip())
    except:
        return 1

def clear_checkpoint():
    if os.path.exists(CHECKPOINT_FILE):
        os.remove(CHECKPOINT_FILE)        

#-------------------------------------     Premium Scrap Extraction Logic     ---------------------------------------------
        
from datetime import datetime
from playwright.async_api import expect

async def policy_premium(page, lob, policy_no, exp_eff_year=2026, exp_exp_year=2027):

    rows = page.locator('[data-automation-id*="vlvwPolicies body-row"]')

    await rows.first.wait_for(timeout=15000)

    count = await rows.count()
    print("Row count:", count)

    for i in range(count):

        row = rows.nth(i)
        cells = row.locator('.body-cell .text')

        try:
            line = (await cells.nth(0).text_content() or "").strip()
            effective = (await cells.nth(2).text_content() or "").strip()
            expiration = (await cells.nth(3).text_content() or "").strip()
            policy_number = (await cells.nth(4).text_content() or "").strip()

            #print(f"Checking Row {i}: LOB={line}, Effective={effective}, Expiration={expiration}")

            print(
                f"Checking Row {i}: "
                f"LOB={line}, "
                f"Policy={policy_number}, "
                f"Effective={effective}, "
                f"Expiration={expiration}"
            )

            if lob.lower() not in line.lower():
                continue

            if policy_no.strip().lower() != policy_number.lower():
                continue

            try:
                eff_year = datetime.strptime(effective, "%m/%d/%Y").year
                exp_year = datetime.strptime(expiration, "%m/%d/%Y").year
            except:
                continue

        except Exception as e:
            print("Row Error:", e)
            continue

        if eff_year == exp_eff_year and exp_year == exp_exp_year:

            print(f"Match found at row {i}, opening...")

            await row.dblclick()

            # =========================
            # NON-WCOM FLOW
            # =========================
            if lob.upper() != "WCOM":

                Commercial_AP = page.locator("text='Commercial AP'")

                await Commercial_AP.wait_for(state="visible", timeout=15000)

                await Commercial_AP.click()

                #await page.locator("text='Commercial AP'").click()

                await page.wait_for_timeout(3000)

                # Click Status
                await page.wait_for_selector("//span[text()='Status']")

                await page.locator("//span[text()='Status']").click()

                # =========================
                # GET PREMIUM
                # =========================

                locator = page.locator('#curePolicyPremium__textField')

                await locator.wait_for(state="attached", timeout=15000)

                # Scroll until visible
                for _ in range(5):

                    await locator.scroll_into_view_if_needed()

                    if await locator.is_visible():
                        break

                    await page.mouse.wheel(0, 1500)

                    await page.wait_for_timeout(1000)

                await locator.wait_for(state="visible", timeout=20000)

                # premium = ""

                # for _ in range(25):

                #     try:
                #         premium = (await locator.input_value()).strip()
                #     except:
                #         premium = (await locator.text_content() or "").strip()

                #     if premium:
                #         break

                #     print("Waiting for premium value...")

                #     await page.wait_for_timeout(1000)

                # if premium and premium != "$":

                #     if not premium.startswith("$"):
                #         premium = f"${premium}"

                #     print("Premium Value:", premium)

                #     return premium

                # return "Premium not found"

                premium = ""

                for _ in range(30):

                    if await locator.is_visible():
                        try:
                            premium = (await locator.input_value()).strip()
                        except:
                            premium = (await locator.text_content() or "").strip()

                        if premium and premium != "$":
                            if not premium.startswith("$"):
                                premium = f"${premium}"

                            print("Premium Value:", premium)
                            return premium

                    print("Waiting for premium value...")
                    await page.wait_for_timeout(1000)

                return "Premium not found"

            # =========================
            # WCOM FLOW
            # =========================
            else:

                premium = ""

                sections = [
                    {
                        "name": "Policy Information/Total Premiums",
                        "locator": '#cureTotalEstimatedAnnualPremium__textField'
                    },
                    {
                        "name": "Total Premium Calculations",
                        "locator": '[data-automation-id="cureTotalAnnualPremium"] input'
                    }
                ]

                for section in sections:

                    section_name = section["name"]

                    locator_value = section["locator"]

                    try:
                        print(f"Opening section: {section_name}")

                        section_tab = page.locator(f"text='{section_name}'")

                        await section_tab.wait_for(state="visible", timeout=15000)

                        await section_tab.click()

                        locator = page.locator(locator_value)

                        # Wait until attached
                        await locator.wait_for(state="attached", timeout=15000)

                        # Scroll until visible
                        for _ in range(5):

                            await locator.scroll_into_view_if_needed()

                            if await locator.is_visible():
                                break

                            await page.mouse.wheel(0, 1500)

                            await page.wait_for_timeout(1000)

                        await locator.wait_for(state="visible", timeout=5000)

                        # for _ in range(15):

                        #     try:
                        #         premium = (await locator.input_value()).strip()
                        #     except:
                        #         premium = (await locator.text_content() or "").strip()

                        #     if premium:
                        #         break

                        #     print("Waiting for premium value...")

                        #     await page.wait_for_timeout(1000)

                        # if premium and premium != "$":

                        #     if not premium.startswith("$"):
                        #         premium = f"${premium}"

                        #     print("Premium Value:", premium)

                        #     return premium

                        premium = ""

                        for _ in range(30):
                            if await locator.is_visible():
                                try:
                                    premium = (await locator.input_value()).strip()
                                except:
                                    premium = (await locator.text_content() or "").strip()

                                if premium and premium != "$":
                                    if not premium.startswith("$"):
                                        premium = f"${premium}"

                                    print("Premium Value:", premium)
                                    return premium

                            print("Waiting for premium value...")
                            await page.wait_for_timeout(1000)

                        print("Premium value not found after 30 seconds.")
                        return None

                    except Exception as e:

                        print(f"Premium not found in section {section}: {e}")

                        continue

                return "Premium not found"

#SCRAPER FUNCTION

async def scrape_records(page, start_id, existing_ids, target_date):

    locator = page.get_by_text("Follow Up/Start")
    await locator.wait_for(state="visible", timeout=20000)

    await page.get_by_text("Customize View").click()
    await page.locator('[data-automation-id="tpgActivitiesDisplayTab"]').click()
    await page.get_by_text("Selected date range").click()

    # yesterday = datetime.now() - timedelta(days=1)
    # yesterday_str = f"{yesterday.month}/{yesterday.day}/{yesterday.year}"
    yesterday_str = f"{target_date.month}/{target_date.day}/{target_date.year}"

    print("Yesterday:", yesterday_str)

    date_inputs = page.locator("input.date-edit[type='text']")
    await date_inputs.first.wait_for(state="visible")

    await date_inputs.nth(0).fill(yesterday_str)
    await date_inputs.nth(1).fill(yesterday_str)

    await page.get_by_role("button", name="Finish").click()

    await page.wait_for_timeout(5000)

    locator = page.get_by_text("Follow Up/Start")
    await locator.wait_for(state="visible", timeout=20000)

    row = load_checkpoint()
    print(f"▶ Resuming from row: {row}")

    current_id = start_id

    MAX_RETRIES = 3

    while True:

        if page.is_closed():
            raise Exception("Page closed → restart browser")

        data_visible = f"/html/body/root/app/program/div/div[2]/div[2]/div/div/screen/div/div/div/proxy[3]/div/asi-panel/div/proxy[3]/div/asi-frame/div/div[2]/proxy/div/asi-frame/div/div[2]/proxy[3]/div/asi-virtual-list-view/div/div/asi-virtual-list-table/div/div/div[2]/div[2]/div/div[{row}]/div[5]/div/span"
        locator = page.locator(f"xpath={data_visible}")

        #await page.wait_for_timeout(5000)

        try:
            await locator.wait_for(state="visible", timeout=10000)
        except:
            print(f"Row {row} not visible — checking scroll...")

            await page.mouse.wheel(0, 2000)
            await page.wait_for_timeout(2000)

            if await locator.count() == 0:
                print("Confirmed end of rows")
                return
            else:
                print("Row loaded after scroll, continuing...")

        await locator.wait_for(state="visible", timeout=20000)

        date_xpath = f"/html/body/root/app/program/div/div[2]/div[2]/div/div/screen/div/div/div/proxy[3]/div/asi-panel/div/proxy[3]/div/asi-frame/div/div[2]/proxy/div/asi-frame/div/div[2]/proxy[3]/div/asi-virtual-list-view/div/div/asi-virtual-list-table/div/div/div[2]/div[2]/div/div[{row}]/div[5]/div/span"

        date_element = page.locator(f"xpath={date_xpath}")
        await date_element.wait_for(timeout=5000)

        date_text = (await date_element.inner_text()).strip()
        print("Date Text:", date_text)
        
        if date_text != yesterday_str:
            row += 1
            continue

        # -------- CREATE REF ID ONCE --------

        timestamp = datetime.now().strftime("%Y%m%d")
        ref_id = f"REF{str(current_id).zfill(6)}_{timestamp}"
        current_id += 1
        
        
        task_folder = os.path.join(PDF_FOLDER, ref_id)
        os.makedirs(task_folder, exist_ok=True)

        retry = 0

        status = "Pending"

        while retry < MAX_RETRIES:

            if page.is_closed():
                raise Exception("Page closed → restart browser")

            try:

                await page.wait_for_timeout(5000)

                print("\n Processing Row:", row, "| Ref ID:", ref_id)

                task_folder = os.path.join(PDF_FOLDER, ref_id)
                os.makedirs(task_folder, exist_ok=True)

                desc_xpath = f"/html/body/root/app/program/div/div[2]/div[2]/div/div/screen/div/div/div/proxy[3]/div/asi-panel/div/proxy[3]/div/asi-frame/div/div[2]/proxy/div/asi-frame/div/div[2]/proxy[3]/div/asi-virtual-list-view/div/div/asi-virtual-list-table/div/div/div[2]/div[2]/div/div[{row}]/div[1]/div/span"

                desc_ele = page.locator(f"xpath={desc_xpath}")
                await desc_ele.wait_for(timeout=10000)
                await desc_ele.scroll_into_view_if_needed()

                row_text = await desc_ele.inner_text()
                print("Row clicked:", row_text)

                await desc_ele.click(force=True)

                await page.wait_for_timeout(10000)

                open_activity = page.get_by_text("Open Activity", exact=True)
                await open_activity.wait_for(state="visible", timeout=20000)

                description = page.locator("//asi-string-edit//div").first
                await description.wait_for(timeout=10000)
                description_value = await description.get_attribute("data-value")
                print("Description:", description_value)

                await page.wait_for_timeout(5000)

                association = page.locator('[data-automation-id="plhAssociation"] .text')
                await association.wait_for(timeout=10000)
                association_value = await association.text_content()
                association_value = association_value.strip() if association_value else ""

                print("Association:", association_value)

                parts = [p.strip() for p in association_value.split(" - ", 2)]

                lob = parts[1] if len(parts) > 1 else ""
                policy_no = parts[2] if len(parts) > 2 else ""

                print("LOB:", lob)
                print("Policy Number:", policy_no)

                # ===== CREATE UNIQUE ID =====

                unique_id = association_value.strip().lower()

                print("CHECK UNIQUE:", unique_id)

                if unique_id in existing_ids:

                    print("Skipping Duplicate:", unique_id)

                    close_btn = page.locator("div.close-button[title*='Close']").last

                    await close_btn.wait_for(state="visible", timeout=10000)

                    await close_btn.click()

                    # await page.locator(
                    #     "div.main-button:has-text('Home')"
                    # ).click()

                    row += 1
                    continue


                await page.locator("text=Account Detail").click()


                account_name = page.locator("//div[text()='Account name']/following::div[@data-value][1]")
                account_name_value = await account_name.get_attribute("data-value")
                print("Account Name:", account_name_value)

                lookup_code = page.locator("//div[text()='Lookup code']/following::div[@data-value][1]")
                lookup_code_value = await lookup_code.get_attribute("data-value")
                print("Lookup Code:", lookup_code_value)

                await page.locator("[data-automation-id='tpgServicingTab']").click()

                account_manager = page.locator('[data-automation-id="cboEmployee1"] input')
                await account_manager.wait_for(timeout=10000)
                account_manager_value = await account_manager.input_value()
                print("Account Manager:", account_manager_value)
                

                #-------------------------------------     Premium & CSR Type Extraction Logic     ---------------------------------------------
                
                await page.locator("//span[normalize-space()='Policies']").click()

                premium = await policy_premium(page, lob, policy_no)
                print("Captured Premium:", premium)

                await page.wait_for_timeout(5000)

                #--------------------------------------------------------------------------------------------------------------------------------

                await page.locator("//span[normalize-space()='Activities']").click(timeout=5000)
                print("Activities clicked")

                await page.wait_for_timeout(5000)

                access = page.locator("div.main-button", has_text="Access")
                await access.wait_for()
                await access.click()
                print("Access clicked to download attachments")
    
                await page.locator("span.text.force-underline", has_text="Attachments").first.click()
                print("Attachments clicked")

                try:
                    await page.wait_for_selector(
                        "//div[contains(@data-automation-id,'vlvwAttachment body-row')]",
                        state="visible",
                        timeout=10000
                    )
                except:
                    await page.wait_for_timeout(20000)


                rows = page.locator("//div[contains(@data-automation-id,'vlvwAttachment body-row')]")
                row_count = await rows.count()

                print("Total Files:", row_count)

                for i in range(row_count):

                    retry = 0
                    while retry < 3:
                        try:

                            file_xpath = f"(//div[contains(@data-automation-id,'vlvwAttachment body-row')]//div[contains(@class,'body-cell first')]//span[contains(@class,'text')])[{i+1}]"
                            file_element = page.locator(f"xpath={file_xpath}")
                            await file_element.wait_for(state="visible", timeout=15000)

                            file_text = (await file_element.inner_text()).strip()
                            print("Processing:", file_text)

                            await file_element.scroll_into_view_if_needed()
                            await file_element.dblclick()

                            download_btn = page.locator("[data-automation-id='Download']")
                            await download_btn.wait_for(state="visible")

                            async with page.expect_download(timeout=20000) as d:
                                await download_btn.click()

                            download = await d.value

                            file_name = re.sub(r'[<>:"/\\|?*]', '-', download.suggested_filename)

                            save_path = os.path.join(task_folder, file_name)
                            await download.save_as(save_path)

                            print("Saved:", save_path)

                            if save_path.lower().endswith(".msg"):
                                extract_pdf_from_msg(save_path)

                            await page.wait_for_timeout(5000)

                            break  

                        except Exception as e:
                            retry += 1
                            print(f"Error at row {i+1} retry {retry}: {e}")
                            await page.wait_for_timeout(3000)
                

                # =========================
                # FINAL STEP: Process MSG after ALL downloads
                # =========================

                target_keywords = [
                    "quote", "binder", "proposal",
                    "schedule", "application", "app", "accord"
                ]

                existing_types = set()

                # Detect OUTER PDFs
                for f in os.listdir(task_folder):
                    f_lower = f.lower()
                    if f_lower.endswith(".pdf"):
                        for k in target_keywords:
                            if k in f_lower:
                                existing_types.add(k)

                
                msg_files = [f for f in os.listdir(task_folder) if f.lower().endswith(".msg")]

                if msg_files:
                    print("Existing OUTER types:", existing_types)

                    for f in msg_files:
                        msg_path = os.path.join(task_folder, f)
                        extract_pdf_from_msg(msg_path, existing_types)

                row_data = {
                    "Date": date_text,
                    "Account Manager": account_manager_value,
                    "Lookup Code": lookup_code_value,
                    "Account Name": account_name_value,
                    "Description": description_value,
                    "Policy number": policy_no,
                    "LOB": lob,
                    "Association": association_value,
                    "Premium": premium,       
                    "CSR Type": "", 
                    "Reference ID": ref_id,
                    "Job ID": "",
                    "Status": "Pending"
                }

                append_to_excel(row_data, existing_ids)
                existing_ids.add(unique_id)
                save_checkpoint(row)

                # await page.locator("div.close-button").wait_for(state="visible")
                # await page.locator("div.close-button").click()

                # await page.locator("div.main-button:has-text('Home')").click()
                # print("Clicked Home button")
                # close_btn = page.locator("div.close-button[title*='Close']").last
                # await close_btn.wait_for(state="visible", timeout=10000)
                # await close_btn.click()

                popup_close = page.locator("button:has([data-icon='SmCloseIcon'])")
                home_btn = page.locator("div.main-button:has-text('Home')")
                close_btn = page.locator("div.close-button[title*='Close']").last

                try:
                    if await popup_close.is_visible(timeout=5000):
                        await popup_close.click()
                        await home_btn.click()
                        print("Popup closed, Home clicked")
                    else:
                        await close_btn.click()
                        print("Close button clicked")
                except Exception:
                    await close_btn.click()
                    print("Close button clicked")

                break

            except Exception as e:
                    
                    await page.wait_for_timeout(10000)

                    retry += 1
                    print(f"\n Error at row {row} retry {retry}")
                    traceback.print_exc()

                    home_btn = page.locator("div.main-button:has-text('Home')")
   
                    if await home_btn.is_visible():
                        await home_btn.click()
                        print("Clicked Home before retry")

                        await page.get_by_text("Follow Up/Start").wait_for(state="visible", timeout=20000)

                    await page.wait_for_timeout(3000)

                    if retry >= MAX_RETRIES:
                        print(f"Skipping row {row} after {MAX_RETRIES} retries")

                        error_data = {
                            "Date": date_text,
                            "Account Manager": account_manager_value,
                            "Lookup Code": lookup_code_value,
                            "Account Name": account_name_value,
                            "Description": description_value,
                            "Policy number": policy_no,
                            "LOB": lob,
                            "Association": association_value,
                            "Premium": "",       
                            "CSR Type": "", 
                            "Reference ID": ref_id,
                            "Job ID": "",
                            "Status": "Error"

                        }

                        append_to_excel(error_data, existing_ids)
                        save_checkpoint(row)
                        break
                
        row += 1

# SAVE TO EXCEL FUNCTION

def append_to_excel(row_data, existing_ids):

    row_data["Reference ID"] = str(row_data["Reference ID"])

    df = pd.DataFrame([row_data])

    if not os.path.exists(EXCEL_FILE_PATH):
        df.to_excel(EXCEL_FILE_PATH, index=False)
        print(f"Saved → {row_data['Reference ID']}\n")
        return True

    from openpyxl import load_workbook
    book = load_workbook(EXCEL_FILE_PATH)
    sheet = book.active
    startrow = sheet.max_row
    book.close()

    with pd.ExcelWriter(
        EXCEL_FILE_PATH,
        engine="openpyxl",
        mode="a",
        if_sheet_exists="overlay"
    ) as writer:

        df.to_excel(
            writer,
            index=False,
            header=False,
            startrow=startrow
        )

    print(f"Saved → {row_data['Reference ID']}\n")
    return True



# MAIN FUNCTION

async def run_epic():

    MAX_BROWSER_RETRIES = 3

    start_id = get_starting_ref_id()

    for attempt in range(MAX_BROWSER_RETRIES):

        browser = None
        context = None
        page = None

        try:
            print(f"\n--------------- EPIC Attempt {attempt+1}")

            existing_ids = set()

            if os.path.exists(EXCEL_FILE_PATH):
                df = pd.read_excel(EXCEL_FILE_PATH)

                df["Date"] = pd.to_datetime(df["Date"], errors="coerce").dt.date
                target_date = (datetime.now() - timedelta(days=2)).date()

                df_filtered = df[df["Date"] == target_date].copy()

                existing_ids = set(
                    df_filtered["Association"]
                    .astype(str)
                    .str.strip()
                    .str.lower()
                )
            else:
                target_date = (datetime.now() - timedelta(days=2)).date()


            async with async_playwright() as p:

                    browser = await p.chromium.launch(channel="chrome", headless=False)
                    context = await browser.new_context()
                    page = await context.new_page()

                    page.set_default_timeout(30000)

                    await login(page, context)
                    await scrape_records(page, start_id, existing_ids, target_date)

                    if not page.is_closed():
                        try:
                            await page.get_by_text("Logout").click(timeout=5000)
                            await page.locator("[data-automation-id='Yes']").click(timeout=5000)
                            await page.wait_for_timeout(3000)
                            print("Logout successful")
                        except Exception as e:
                            print("Logout failed:", e)

                    print("EPIC completed successfully")
                    clear_checkpoint()
                    return 

        # except Exception as e:
        #     print(f"\n Browser failed: {e}")
        #     await asyncio.sleep(5)

        except PlaywrightError as e:
            error_text = str(e)

            print(f"\n Playwright error: {error_text}")

            if "Target closed" in error_text:
                print("Detected browser crash → restarting...")

            await asyncio.sleep(5)

            
        # finally:
        #     print("Cleaning EPIC resources...")

        #     if page:
        #         try:
        #             await page.title()
        #         except:
        #             print("Page already closed, skipping logout")

        #     try:
        #         if context:
        #             await context.close()
        #         if browser:
        #             await browser.close()
        #     except:
        #         pass

        finally:
            print("Cleaning EPIC resources...")

            if page:
                if page.is_closed():
                    print("Page already closed")
                else:
                    print("Page still open")

            try:
                if context:
                    await context.close()
                if browser:
                    await browser.close()
            except Exception as e:
                print("Cleanup error:", e)

    print("EPIC failed after max retries")

if __name__ == "__main__":
    asyncio.run(run_epic())

