# ##########################################################  Correcting Unique ID Logic with Association IDs ##########################################################        CORRECT AND WORKING CODE

# ##########------------##########------------#########        GOOD FOR DAILY RUNS


# import asyncio
# from playwright.async_api import async_playwright
# from configparser import ConfigParser
# from datetime import datetime, timedelta
# import pandas as pd
# import os
# import re
# import extract_msg
# from openpyxl.utils import get_column_letter
# import traceback
# from playwright.async_api import Error as PlaywrightError
# from playwright.async_api import TimeoutError

# config = ConfigParser()
# config.read("c:/Users/vijay_m/OneDrive - Exdion Solutions Pvt. Ltd.-70692290/Documents/Project_Job_Creation/Carroll/config.ini")
# BASE_FOLDER = config.get("PATHS", "BASE_FOLDER")
# PDF_FOLDER = os.path.join(BASE_FOLDER, "Job_Creation")
# EXCEL_FILE_PATH = os.path.join(BASE_FOLDER, "Carroll_Tracker.xlsx")
# CHECKPOINT_FILE = os.path.join(BASE_FOLDER, "epic_checkpoint.txt")

# os.makedirs(BASE_FOLDER, exist_ok=True)
# os.makedirs(PDF_FOLDER, exist_ok=True)

# # ================= GET STARTING REFERENCE NUMBER =================

# def get_starting_ref_id():

#     if not os.path.exists(EXCEL_FILE_PATH):
#         return 1

#     df = pd.read_excel(EXCEL_FILE_PATH)

#     if "Reference ID" not in df.columns or df.empty:
#         return 1

#     existing_ids = df["Reference ID"].dropna()

#     numbers = (
#         existing_ids
#         .str.extract(r'REF(\d+)')[0]
#         .dropna()
#         .astype(int)
#     )

#     return numbers.max() + 1 if not numbers.empty else 1


# # LOGIN FUNCTION

# async def login(page, context):

#     url = config.get("CarrollEpic", "Link")
#     EnterpriseID = config.get("CarrollEpic", "Enterprise_ID")
#     username = config.get("CarrollEpic", "User_ID")
#     password = config.get("CarrollEpic", "Password")

#     await page.goto(url, timeout=90000)

#     await page.get_by_role("textbox").fill(EnterpriseID) 

#     await page.locator("//button[.//span[contains(text(),'Contin')]]").click()

#     async with context.expect_page() as new_page_event:
#         await page.locator("//button[contains(text(),'Login')]").click()

#     login_page = await new_page_event.value

#     await login_page.wait_for_load_state("networkidle")

#     await login_page.locator("#usercode").fill(username)
#     await login_page.locator("#password").fill(password)

#     await login_page.locator("//button[.//div[text()='Login']]").click()

#     try:
#         later_btn = login_page.get_by_role("button", name="Later")
#         await later_btn.wait_for(timeout=5000)
#         await later_btn.click()
#         print("Later Reset clicked")
#     except:
#         pass
    
#     await page.wait_for_load_state("networkidle")

#     await page.locator("//div[contains(@class,'drop-btn')]").first.click()
#     await page.locator("//span[@class='text' and text()='CARRO09_PROD']").click()

#     await page.locator("//button[.//span[contains(text(),'Contin')]]").click()

#     try:
#         await page.locator("//button[@data-automation-id='Yes']").click(timeout=3000)
#         print("Popup YES clicked, Login successful")
#     except:
#         print("EPIC Login Successful")

#     #await page.pause()

#     return login_page


# # ------------- CODE FOR DOWNLOAD .MSG FILE AT LAST----------------

# def extract_pdf_from_msg(msg_path, existing_types):
#     try:
#         msg = extract_msg.Message(msg_path)
#         save_folder = os.path.dirname(msg_path)

#         target_keywords = [
#             "quote", "binder", "proposal",
#             "schedule", "application", "app", "accord"
#         ]

#         for att in msg.attachments:
#             filename = att.longFilename or att.shortFilename
#             if not filename:
#                 continue

#             filename_lower = filename.lower()

#             if filename_lower.endswith(".pdf"):

#                 matched_keyword = next((k for k in target_keywords if k in filename_lower), None)

#                 if matched_keyword:
#                     if matched_keyword in existing_types:
#                         print(f"Skipping INNER {matched_keyword}:", filename)
#                         continue

#                     pdf_path = os.path.join(save_folder, filename)

#                     with open(pdf_path, "wb") as f:
#                         f.write(att.data)

#                     print("Saved INNER:", pdf_path)

#                     existing_types.add(matched_keyword)

#     except Exception as e:
#         print("MSG Extraction Error:", e)

# #--------------------Checkpoint functions to save progress and avoid duplicates in case of crash-------------------#

# def save_checkpoint(row):
#     with open(CHECKPOINT_FILE, "w") as f:
#         f.write(str(row))

# def load_checkpoint():
#     if not os.path.exists(CHECKPOINT_FILE):
#         return 1
#     try:
#         with open(CHECKPOINT_FILE, "r") as f:
#             return int(f.read().strip())
#     except:
#         return 1

# def clear_checkpoint():
#     if os.path.exists(CHECKPOINT_FILE):
#         os.remove(CHECKPOINT_FILE)        
# #------------------------------------------------------#

# # SCRAPER FUNCTION

# async def scrape_records(page, start_id, existing_ids, target_date):

#     locator = page.get_by_text("Follow Up/Start")
#     await locator.wait_for(state="visible", timeout=20000)

#     await page.get_by_text("Customize View").click()
#     await page.locator('[data-automation-id="tpgActivitiesDisplayTab"]').click()
#     await page.get_by_text("Selected date range").click()

#     # yesterday = datetime.now() - timedelta(days=1)
#     # yesterday_str = f"{yesterday.month}/{yesterday.day}/{yesterday.year}"
#     yesterday_str = f"{target_date.month}/{target_date.day}/{target_date.year}"

#     print("Yesterday:", yesterday_str)

#     date_inputs = page.locator("input.date-edit[type='text']")
#     await date_inputs.first.wait_for(state="visible")

#     await date_inputs.nth(0).fill(yesterday_str)
#     await date_inputs.nth(1).fill(yesterday_str)

#     await page.get_by_role("button", name="Finish").click()

#     await page.wait_for_timeout(5000)

#     locator = page.get_by_text("Follow Up/Start")
#     await locator.wait_for(state="visible", timeout=20000)

#     row = load_checkpoint()
#     print(f"▶ Resuming from row: {row}")

#     current_id = start_id

#     MAX_RETRIES = 3

#     while True:
        
#         if page.is_closed():
#             raise Exception("Page closed → restart browser")

#         data_visible = f"/html/body/root/app/program/div/div[2]/div[2]/div/div/screen/div/div/div/proxy[3]/div/asi-panel/div/proxy[3]/div/asi-frame/div/div[2]/proxy/div/asi-frame/div/div[2]/proxy[3]/div/asi-virtual-list-view/div/div/asi-virtual-list-table/div/div/div[2]/div[2]/div/div[{row}]/div[4]/div/span"
#         locator = page.locator(f"xpath={data_visible}")

#         #await page.wait_for_timeout(5000)

#         try:
#             await locator.wait_for(state="visible", timeout=10000)
#         except:
#             print(f"Row {row} not visible — checking scroll...")

#             await page.mouse.wheel(0, 2000)
#             await page.wait_for_timeout(2000)

#             if await locator.count() == 0:
#                 print("Confirmed end of rows")
#                 return
#             else:
#                 print("Row loaded after scroll, continuing...")
        
#         await locator.wait_for(state="visible", timeout=20000)

#         date_xpath = f"/html/body/root/app/program/div/div[2]/div[2]/div/div/screen/div/div/div/proxy[3]/div/asi-panel/div/proxy[3]/div/asi-frame/div/div[2]/proxy/div/asi-frame/div/div[2]/proxy[3]/div/asi-virtual-list-view/div/div/asi-virtual-list-table/div/div/div[2]/div[2]/div/div[{row}]/div[4]/div/span"

#         date_element = page.locator(f"xpath={date_xpath}")
#         await date_element.wait_for(timeout=5000)

#         date_text = (await date_element.inner_text()).strip()
#         print(f"Date Text: {date_text}")
        
#         if date_text != yesterday_str:
#             row += 1
#             continue

#         retry = 0

#         status = "Pending"

#         duplicate_found = False

#         while retry < MAX_RETRIES:

#             if page.is_closed():
#                 raise Exception("Page closed → restart browser")

#             try:

#                 await page.wait_for_timeout(5000)
#                 print(f"\n========== Processing Row {row} ==========")
                               
#                 desc_xpath = f"/html/body/root/app/program/div/div[2]/div[2]/div/div/screen/div/div/div/proxy[3]/div/asi-panel/div/proxy[3]/div/asi-frame/div/div[2]/proxy/div/asi-frame/div/div[2]/proxy[3]/div/asi-virtual-list-view/div/div/asi-virtual-list-table/div/div/div[2]/div[2]/div/div[{row}]/div[1]/div/span"

#                 desc_ele = page.locator(f"xpath={desc_xpath}")

#                 try:
#                     await desc_ele.wait_for(state="visible", timeout=10000)
#                 except:
#                     print(f"Row {row} not found, skipping")
#                     retry = MAX_RETRIES
#                     break

#                 await desc_ele.scroll_into_view_if_needed()

#                 row_text = (await desc_ele.inner_text()).strip()
#                 print(f"Clicking Row {row}: {row_text}")
                
#                 await desc_ele.click()

#                 await desc_ele.click(force=True)
#                 print(f"Row {row} clicked")

#                 # await desc_ele.click()
#                 # await page.wait_for_timeout(300)
#                 # await desc_ele.click()

#                 # await page.wait_for_timeout(10000)

#                 open_activity = page.get_by_text("Open Activity", exact=True)
#                 await open_activity.wait_for(state="visible", timeout=40000)
                
#                 try:
#                     description = page.locator("//asi-string-edit//div").first
#                     await description.wait_for(timeout=10000)
#                     description_value = await description.get_attribute("data-value")
#                 except TimeoutError:
#                     description = page.locator("//input[@id='streDescription']")
#                     await description.wait_for(state="visible", timeout=10000)
#                     description_value = await description.input_value()

#                 print("Description:", description_value)

#                 await page.wait_for_timeout(5000)

#                 try:
#                     association = page.locator('[data-automation-id="plhAssociation"] .text')
#                     await association.wait_for(timeout=5000)

#                 except TimeoutError:
#                     association = page.locator("//dt[normalize-space()='Association']/following-sibling::div/dd[1]/div")
#                     await association.wait_for(timeout=5000)

#                 association_value = (await association.text_content() or "").strip()
#                 print("Association:", association_value)

#                 parts = [p.strip() for p in association_value.split(" - ", 2)]

#                 lob = parts[1] if len(parts) > 1 else ""
#                 policy_no = parts[2] if len(parts) > 2 else ""

#                 print("LOB:", lob)
#                 print("Policy Number:", policy_no)

#                 # ===== CREATE UNIQUE ID =====

#                 unique_id = association_value.strip().lower()

#                 print("CHECK UNIQUE:", unique_id)

#                 if unique_id in existing_ids:

#                     print("Skipping Duplicate:", unique_id)

#                     try:
#                         await page.locator("div.close-button").first.click(timeout=3000)
#                     except:
#                         pass

#                     duplicate_found = True
#                     break

#                 timestamp = datetime.now().strftime("%Y%m%d")
#                 ref_id = f"REF{str(current_id).zfill(6)}_{timestamp}"
#                 current_id += 1

#                 print(f"Processing Row {row}: | Ref ID: {ref_id}")

#                 task_folder = os.path.join(PDF_FOLDER, ref_id)
#                 os.makedirs(task_folder, exist_ok=True)

#                 await page.locator("text=Account Detail").click()

#                 account_name = page.locator("//div[text()='Account name']/following::div[@data-value][1]")
#                 account_name_value = await account_name.get_attribute("data-value")
#                 print("Account Name:", account_name_value)

#                 lookup_code = page.locator("//div[text()='Lookup code']/following::div[@data-value][1]")
#                 lookup_code_value = await lookup_code.get_attribute("data-value")
#                 print("Lookup Code:", lookup_code_value)

#                 await page.locator("[data-automation-id='tpgServicingTab']").click()

#                 account_manager = page.locator('[data-automation-id="cboEmployee1"] input')
#                 await account_manager.wait_for(timeout=10000)
#                 account_manager_value = await account_manager.input_value()
#                 print("Account Manager:", account_manager_value)

#                 await page.locator("//span[normalize-space()='Activities']").click(timeout=5000)
#                 print("Activities clicked")

#                 await page.wait_for_timeout(5000)

#                 access = page.locator("div.main-button", has_text="Access")
#                 await access.wait_for()
#                 await access.click()
#                 print("Access clicked to download attachments")
    
#                 await page.locator("span.text.force-underline", has_text="Attachments").first.click()
#                 print("Attachments clicked")

#                 try:
#                     await page.wait_for_selector(
#                         "//div[contains(@data-automation-id,'vlvwAttachment body-row')]",
#                         state="visible",
#                         timeout=10000
#                     )
#                 except:
#                     await page.wait_for_timeout(20000)

#                 rows = page.locator("//div[contains(@data-automation-id,'vlvwAttachment body-row')]")
#                 row_count = await rows.count()

#                 print("Total Files:", row_count)

#                 for i in range(row_count):

#                     retry = 0
#                     while retry < 3:
#                         try:

#                             file_xpath = f"(//div[contains(@data-automation-id,'vlvwAttachment body-row')]//div[contains(@class,'body-cell first')]//span[contains(@class,'text')])[{i+1}]"
#                             file_element = page.locator(f"xpath={file_xpath}")
#                             await file_element.wait_for(state="visible", timeout=15000)

#                             file_text = (await file_element.inner_text()).strip()
#                             print("Processing:", file_text)

#                             await file_element.scroll_into_view_if_needed()
#                             await file_element.dblclick()

#                             download_btn = page.locator("[data-automation-id='Download']")
#                             await download_btn.wait_for(state="visible")

#                             async with page.expect_download(timeout=20000) as d:
#                                 await download_btn.click()

#                             download = await d.value

#                             file_name = re.sub(r'[<>:"/\\|?*]', '-', download.suggested_filename)

#                             save_path = os.path.join(task_folder, file_name)
#                             await download.save_as(save_path)

#                             print("Saved:", save_path)

#                             await page.wait_for_timeout(3000)

#                             break

#                         except Exception as e:
#                             retry += 1
#                             print(f"Error at row {i+1} retry {retry}: {e}")
#                             await page.wait_for_timeout(3000)


#                 # =========================
#                 # FINAL STEP: Process MSG after ALL downloads
#                 # =========================

#                 target_keywords = [
#                     "quote", "binder", "proposal",
#                     "schedule", "application", "app", "accord"
#                 ]

#                 existing_types = set()

#                 # Detect OUTER PDFs
#                 for f in os.listdir(task_folder):
#                     f_lower = f.lower()
#                     if f_lower.endswith(".pdf"):
#                         for k in target_keywords:
#                             if k in f_lower:
#                                 existing_types.add(k)

#                 # print("Existing OUTER types:", existing_types)

#                 # # Process MSG files AFTER
#                 # for f in os.listdir(task_folder):
#                 #     if f.lower().endswith(".msg"):
#                 #         msg_path = os.path.join(task_folder, f)
#                 #         extract_pdf_from_msg(msg_path, existing_types)

#                 msg_files = [f for f in os.listdir(task_folder) if f.lower().endswith(".msg")]

#                 if msg_files:
#                     print("Existing OUTER types:", existing_types)

#                     for f in msg_files:
#                         msg_path = os.path.join(task_folder, f)
#                         extract_pdf_from_msg(msg_path, existing_types)


#                 row_data = {
#                     "Date": date_text,
#                     "Account Manager": account_manager_value,
#                     "Lookup Code": lookup_code_value,
#                     "Account Name": account_name_value,
#                     "Description": description_value,
#                     "Policy number": policy_no,
#                     "LOB": lob,
#                     "Association": association_value,
#                     "Job ID": "",
#                     "Status": "Pending",
#                     "Reference ID": ref_id
#                 }

#                 append_to_excel(row_data, existing_ids)
#                 existing_ids.add(unique_id)
#                 save_checkpoint(row)

#                 # await page.locator("div.close-button").wait_for(state="visible")
#                 # await page.locator("div.close-button").click()

#                 # # close_btn = page.locator("div.close-button[title*='Close']").last
#                 # # await close_btn.wait_for(state="visible", timeout=10000)
#                 # # await close_btn.click()

#                 home_btn = page.locator("div.main-button:has-text('Home')")
#                 if await home_btn.is_visible():
#                     await home_btn.click()
#                     print("Clicked home button")

#                     await page.get_by_text("Follow Up/Start").wait_for(state="visible", timeout=20000)

#                 await page.wait_for_timeout(3000)

#                 print("\n-------------------------------------------------------------------------------------------------------------------------------------------------------------------\n")

#                 break

#             except Exception as e:
                    
#                     await page.wait_for_timeout(10000)

#                     retry += 1
#                     #print(f"Error at row {row} retry {retry}: {e}")
#                     print(f"\n ERROR at row {row} retry {retry}")
#                     traceback.print_exc()

#                     home_btn = page.locator("div.main-button:has-text('Home')")

#                     if await home_btn.is_visible():
#                         await home_btn.click()
#                         print("Clicked Home before retry")

#                         await page.get_by_text("Follow Up/Start").wait_for(state="visible", timeout=20000)

#                     await page.wait_for_timeout(3000)

#                     if retry >= MAX_RETRIES:
#                         print(f"Skipping row {row} after {MAX_RETRIES} retries")

#                         error_data = {
#                             "Date": date_text,
#                             "Account Manager": account_manager_value,
#                             "Lookup Code": lookup_code_value,
#                             "Account Name": account_name_value,
#                             "Description": description_value,
#                             "Policy number": policy_no,
#                             "LOB": lob,
#                             "Association": association_value,
#                             "Job ID": "",
#                             "Status": "Error",
#                             "Reference ID": ref_id
#                         }

#                         append_to_excel(error_data, existing_ids)
#                         save_checkpoint(row)
#                         break

#         if duplicate_found:
#             row += 1
#             continue

#         row += 1

# # SAVE TO EXCEL FUNCTION

# def append_to_excel(row_data, existing_ids):

#     row_data["Reference ID"] = str(row_data["Reference ID"])

#     df = pd.DataFrame([row_data])

#     if not os.path.exists(EXCEL_FILE_PATH):
#         df.to_excel(EXCEL_FILE_PATH, index=False)
#         print(f"Saved → {row_data['Reference ID']}")
#         return True

#     from openpyxl import load_workbook
#     book = load_workbook(EXCEL_FILE_PATH)
#     sheet = book.active
#     startrow = sheet.max_row
#     book.close()

#     with pd.ExcelWriter(
#         EXCEL_FILE_PATH,
#         engine="openpyxl",
#         mode="a",
#         if_sheet_exists="overlay"
#     ) as writer:

#         df.to_excel(
#             writer,
#             index=False,
#             header=False,
#             startrow=startrow
#         )

#     print(f"Saved → {row_data['Reference ID']}")
#     return True

# # MAIN FUNCTION

# async def run_epic():

#     MAX_BROWSER_RETRIES = 3

#     start_id = get_starting_ref_id()

#     for attempt in range(MAX_BROWSER_RETRIES):

#         browser = None
#         context = None
#         page = None

#         try:
#             print(f"\n--------------- EPIC Attempt {attempt+1}")

#             existing_ids = set()

#             if os.path.exists(EXCEL_FILE_PATH):
#                 df = pd.read_excel(EXCEL_FILE_PATH)

#                 df["Date"] = pd.to_datetime(df["Date"], errors="coerce").dt.date
#                 target_date = (datetime.now() - timedelta(days=2)).date()

#                 df_filtered = df[df["Date"] == target_date].copy()

#                 existing_ids = set(
#                     df_filtered["Association"]
#                     .astype(str)
#                     .str.strip()
#                     .str.lower()
#                 )
#             else:
#                 target_date = (datetime.now() - timedelta(days=2)).date()

#             async with async_playwright() as p:

#                 browser = await p.chromium.launch(channel="chrome", headless=False)
#                 context = await browser.new_context()
#                 page = await context.new_page()

#                 page.set_default_timeout(30000)

#                 await login(page, context)
#                 await scrape_records(page, start_id, existing_ids, target_date)

#                 if not page.is_closed():
#                     try:
#                         await page.get_by_text("Logout").click(timeout=5000)
#                         await page.locator("[data-automation-id='Yes']").click(timeout=5000)
#                         await page.wait_for_timeout(3000)
#                         print("Logout successful")
#                     except Exception as e:
#                         print("Logout failed:", e)

#                 print("EPIC completed successfully")
#                 clear_checkpoint()
#                 return 
            
#         # except Exception as e:
#         #     print(f"\n Browser failed: {e}")
#         #     await asyncio.sleep(5)

#         except PlaywrightError as e:
#             error_text = str(e)

#             print(f"\n Playwright error: {error_text}")

#             if "Target closed" in error_text:
#                 print("Detected browser crash → restarting...")

#             await asyncio.sleep(5)

#         # finally:
#         #     print("Cleaning EPIC resources...")

#         #     try:
#         #         if browser:
#         #             await browser.close()
#         #     except Exception as e:
#         #         print("Browser cleanup error:", e)

#         finally:
#             print("Cleaning EPIC resources...")

#             try:
#                 if browser and browser.is_connected():
#                     await browser.close()
#                     print("Browser closed successfully")
#             except Exception as e:
#                 print("Browser cleanup error:", e)

#     print("EPIC failed after max retries")


# if __name__ == "__main__":
#     asyncio.run(run_epic())                                           



##########################################################  Date Specification & Logic Changes ##############################################################

import asyncio
from playwright.async_api import async_playwright
from configparser import ConfigParser
from datetime import datetime, timedelta
import pandas as pd
import os
import re
import extract_msg
from openpyxl.utils import get_column_letter
import traceback
from playwright.async_api import Error as PlaywrightError
from playwright.async_api import TimeoutError

config = ConfigParser()
config.read("c:/Users/vijay_m/OneDrive - Exdion Solutions Pvt. Ltd.-70692290/Documents/Project_Job_Creation/Carroll/config.ini")
BASE_FOLDER = config.get("PATHS", "BASE_FOLDER")
PDF_FOLDER = os.path.join(BASE_FOLDER, "Job_Creation")
EXCEL_FILE_PATH = os.path.join(BASE_FOLDER, "Carroll_Tracker.xlsx")
CHECKPOINT_FILE = os.path.join(BASE_FOLDER, "epic_checkpoint.txt")

os.makedirs(BASE_FOLDER, exist_ok=True)
os.makedirs(PDF_FOLDER, exist_ok=True)

# ================= GET STARTING REFERENCE NUMBER =================

def get_starting_ref_id():

    if not os.path.exists(EXCEL_FILE_PATH):
        return 1

    df = pd.read_excel(EXCEL_FILE_PATH)

    if "Reference ID" not in df.columns or df.empty:
        return 1

    existing_ids = df["Reference ID"].dropna()

    numbers = (
        existing_ids
        .str.extract(r'REF(\d+)')[0]
        .dropna()
        .astype(int)
    )

    return numbers.max() + 1 if not numbers.empty else 1


# LOGIN FUNCTION

# async def login(page, context):

#     url = config.get("CarrollEpic", "Link")
#     EnterpriseID = config.get("CarrollEpic", "Enterprise_ID")
#     username = config.get("CarrollEpic", "User_ID")
#     password = config.get("CarrollEpic", "Password")

#     await page.goto(url, timeout=90000)

#     await page.get_by_role("textbox").fill(EnterpriseID) 

#     await page.locator("//button[.//span[contains(text(),'Contin')]]").click()

#     async with context.expect_page() as new_page_event:
#         await page.locator("//button[contains(text(),'Login')]").click()

#     login_page = await new_page_event.value

#     await login_page.locator("#usercode").wait_for(state="visible",timeout=60000)

#     await login_page.locator("#usercode").fill(username)
#     await login_page.locator("#password").fill(password)

#     await login_page.locator("//button[.//div[text()='Login']]").click()
    
#     try:
#         later_btn = login_page.get_by_role("button", name="Later")
#         await later_btn.wait_for(timeout=5000)
#         await later_btn.click()
#         print("Later Reset clicked")
#     except:
#         pass

#     dropdown = page.locator("//div[contains(@class,'drop-btn')]").first

#     await dropdown.wait_for(state="visible", timeout=90000)
#     await dropdown.click()

#     await page.locator("//span[@class='text' and text()='CARRO09_PROD']").click()

#     await page.locator("//button[.//span[contains(text(),'Contin')]]").click()

#     try:
#         await page.locator("//button[@data-automation-id='Yes']").click(timeout=3000)
#         print("Popup YES clicked, Login successful")
#     except:
#         print("EPIC Login Successful")

#     #await page.pause()

#     return login_page


async def login(page, context):

    url = config.get("CarrollEpic", "Link")
    enterprise_id = config.get("CarrollEpic", "Enterprise_ID")
    username = config.get("CarrollEpic", "User_ID")
    password = config.get("CarrollEpic", "Password")

    print("Opening EPIC...")

    await page.goto(url,wait_until="domcontentloaded",timeout=90000)

    await page.get_by_role("textbox").fill(enterprise_id)

    await page.locator("//button[.//span[contains(text(),'Contin')]]").click()

    print("Opening login popup...")

    async with context.expect_page(timeout=90000) as popup:
        await page.locator("//button[contains(text(),'Login')]").click()

    login_page = await popup.value

    await login_page.wait_for_load_state("domcontentloaded")

    for attempt in range(3):
        try:
            print(f"Waiting for login page... Attempt {attempt + 1}")

            await login_page.wait_for_url("**/EpicIdentityProvider/**",timeout=90000)

            await login_page.locator("#usercode").wait_for(state="visible",timeout=30000)

            print("Login page loaded.")
            break

        except TimeoutError:

            if attempt == 2:
                raise

            print("Retrying login page...")
            await asyncio.sleep(2)

            try:
                await login_page.reload(wait_until="domcontentloaded",timeout=90000)
            except Exception:
                pass

    await login_page.locator("#usercode").fill(username)
    await login_page.locator("#password").fill(password)

    print("Click Login")

    await login_page.locator("//button[.//div[text()='Login']]").click()

    try:
        await login_page.wait_for_event("close", timeout=120000)
        print("Login popup closed")
    except TimeoutError:
        print("Popup still open")

    dropdown = page.locator(".drop-btn").first

    await dropdown.wait_for(state="visible",timeout=120000)

    print("EPIC Home loaded")

    await dropdown.click()

    await page.locator("//span[text()='CARRO09_PROD']").click()

    await page.locator("//button[.//span[contains(text(),'Contin')]]").click()

    try:
        yes_btn = page.locator("//button[@data-automation-id='Yes']")
        await yes_btn.wait_for(state="visible",timeout=5000)
        await yes_btn.click()
        print("YES clicked")
    except TimeoutError:
        pass

    print("EPIC Login Successful")

    return page



# ------------- CODE FOR DOWNLOAD .MSG FILE AT LAST----------------

def extract_pdf_from_msg(msg_path, existing_types):
    try:
        msg = extract_msg.Message(msg_path)
        save_folder = os.path.dirname(msg_path)

        target_keywords = [
            "quote", "binder", "proposal",
            "schedule", "application", "app", "accord"
        ]

        for att in msg.attachments:
            filename = att.longFilename or att.shortFilename
            if not filename:
                continue

            filename_lower = filename.lower()

            if filename_lower.endswith(".pdf"):

                matched_keyword = next((k for k in target_keywords if k in filename_lower), None)

                if matched_keyword:
                    if matched_keyword in existing_types:
                        print(f"Skipping INNER {matched_keyword}:", filename)
                        continue

                    pdf_path = os.path.join(save_folder, filename)

                    with open(pdf_path, "wb") as f:
                        f.write(att.data)

                    print("Saved INNER:", pdf_path)

                    existing_types.add(matched_keyword)

    except Exception as e:
        print("MSG Extraction Error:", e)

#--------------------Checkpoint functions to save progress and avoid duplicates in case of crash-------------------#

def save_checkpoint(target_date, row):
    with open(CHECKPOINT_FILE, "w") as f:
        f.write(f"{target_date},{row}")


def load_checkpoint(target_date):
    if not os.path.exists(CHECKPOINT_FILE):
        return 1

    try:
        with open(CHECKPOINT_FILE, "r") as f:
            saved_date, saved_row = f.read().strip().split(",")

        if saved_date == str(target_date):
            return int(saved_row)

        return 1

    except:
        return 1


def clear_checkpoint():
    if os.path.exists(CHECKPOINT_FILE):
        os.remove(CHECKPOINT_FILE)     
#------------------------------------------------------#

# SCRAPER FUNCTION

async def scrape_records(page, start_id, existing_ids, target_date):

    locator = page.get_by_text("Follow Up/Start")
    await locator.wait_for(state="visible", timeout=20000)

    await page.get_by_text("Customize View").click()
    await page.locator('[data-automation-id="tpgActivitiesDisplayTab"]').click()
    await page.get_by_text("Selected date range").click()

    # yesterday = datetime.now() - timedelta(days=1)
    # yesterday_str = f"{yesterday.month}/{yesterday.day}/{yesterday.year}"
    yesterday_str = f"{target_date.month}/{target_date.day}/{target_date.year}"

    print("Yesterday:", yesterday_str)

    date_inputs = page.locator("input.date-edit[type='text']")
    await date_inputs.first.wait_for(state="visible")

    await date_inputs.nth(0).fill(yesterday_str)
    await date_inputs.nth(1).fill(yesterday_str)

    await page.get_by_role("button", name="Finish").click()

    await page.wait_for_timeout(5000)

    locator = page.get_by_text("Follow Up/Start")
    await locator.wait_for(state="visible", timeout=20000)

    row = load_checkpoint(target_date)
    print(f"▶ Resuming from row: {row}")

    current_id = start_id

    MAX_RETRIES = 3

    while True:
        
        if page.is_closed():
            raise Exception("Page closed → restart browser")

        data_visible = f"/html/body/root/app/program/div/div[2]/div[2]/div/div/screen/div/div/div/proxy[3]/div/asi-panel/div/proxy[3]/div/asi-frame/div/div[2]/proxy/div/asi-frame/div/div[2]/proxy[3]/div/asi-virtual-list-view/div/div/asi-virtual-list-table/div/div/div[2]/div[2]/div/div[{row}]/div[4]/div/span"
        locator = page.locator(f"xpath={data_visible}")

        #await page.wait_for_timeout(5000)

        try:
            await locator.wait_for(state="visible", timeout=10000)
        except:
            print(f"Row {row} not visible — checking scroll...")

            await page.mouse.wheel(0, 2000)
            await page.wait_for_timeout(2000)

            if await locator.count() == 0:
                print("Confirmed end of rows")
                return
            else:
                print("Row loaded after scroll, continuing...")
        
        await locator.wait_for(state="visible", timeout=20000)

        date_xpath = f"/html/body/root/app/program/div/div[2]/div[2]/div/div/screen/div/div/div/proxy[3]/div/asi-panel/div/proxy[3]/div/asi-frame/div/div[2]/proxy/div/asi-frame/div/div[2]/proxy[3]/div/asi-virtual-list-view/div/div/asi-virtual-list-table/div/div/div[2]/div[2]/div/div[{row}]/div[4]/div/span"

        date_element = page.locator(f"xpath={date_xpath}")
        await date_element.wait_for(timeout=5000)

        date_text = (await date_element.inner_text()).strip()
        print(f"Date Text: {date_text}")
        
        if date_text != yesterday_str:
            row += 1
            continue

        retry = 0

        status = "Pending"

        duplicate_found = False

        while retry < MAX_RETRIES:

            if page.is_closed():
                raise Exception("Page closed → restart browser")

            try:

                await page.wait_for_timeout(5000)
                print(f"\n========== Processing Row {row} ==========")
                               
                desc_xpath = f"/html/body/root/app/program/div/div[2]/div[2]/div/div/screen/div/div/div/proxy[3]/div/asi-panel/div/proxy[3]/div/asi-frame/div/div[2]/proxy/div/asi-frame/div/div[2]/proxy[3]/div/asi-virtual-list-view/div/div/asi-virtual-list-table/div/div/div[2]/div[2]/div/div[{row}]/div[1]/div/span"

                desc_ele = page.locator(f"xpath={desc_xpath}")

                try:
                    await desc_ele.wait_for(state="visible", timeout=10000)
                except:
                    print(f"Row {row} not found, skipping")
                    retry = MAX_RETRIES
                    break

                await desc_ele.scroll_into_view_if_needed()

                row_text = (await desc_ele.inner_text()).strip()
                print(f"Clicking Row {row}: {row_text}")
                
                await desc_ele.click()

                await desc_ele.click(force=True)
                print(f"Row {row} clicked")

                # await desc_ele.click()
                # await page.wait_for_timeout(300)
                # await desc_ele.click()

                # await page.wait_for_timeout(10000)

                open_activity = page.get_by_text("Open Activity", exact=True)
                await open_activity.wait_for(state="visible", timeout=40000)
                
                try:
                    description = page.locator("//asi-string-edit//div").first
                    await description.wait_for(timeout=10000)
                    description_value = await description.get_attribute("data-value")
                except TimeoutError:
                    description = page.locator("//input[@id='streDescription']")
                    await description.wait_for(state="visible", timeout=10000)
                    description_value = await description.input_value()

                print("Description:", description_value)

                await page.wait_for_timeout(5000)

                try:
                    association = page.locator('[data-automation-id="plhAssociation"] .text')
                    await association.wait_for(timeout=5000)

                except TimeoutError:
                    association = page.locator("//dt[normalize-space()='Association']/following-sibling::div/dd[1]/div")
                    await association.wait_for(timeout=5000)

                association_value = (await association.text_content() or "").strip()
                print("Association:", association_value)

                parts = [p.strip() for p in association_value.split(" - ", 2)]

                lob = parts[1] if len(parts) > 1 else ""
                policy_no = parts[2] if len(parts) > 2 else ""

                print("LOB:", lob)
                print("Policy Number:", policy_no)

                # ===== CREATE UNIQUE ID =====

                unique_id = association_value.strip().lower()

                print("CHECK UNIQUE:", unique_id)

                if unique_id in existing_ids:

                    print("Skipping Duplicate:", unique_id)

                    try:
                        await page.locator("div.close-button").first.click(timeout=3000)
                    except:
                        pass

                    duplicate_found = True
                    break

                timestamp = datetime.now().strftime("%Y%m%d")
                ref_id = f"REF{str(current_id).zfill(6)}_{timestamp}"
                current_id += 1

                print(f"Processing Row {row}: | Ref ID: {ref_id}")

                task_folder = os.path.join(PDF_FOLDER, ref_id)
                os.makedirs(task_folder, exist_ok=True)

                await page.locator("text=Account Detail").click()

                account_name = page.locator("//div[text()='Account name']/following::div[@data-value][1]")
                account_name_value = await account_name.get_attribute("data-value")
                print("Account Name:", account_name_value)

                lookup_code = page.locator("//div[text()='Lookup code']/following::div[@data-value][1]")
                lookup_code_value = await lookup_code.get_attribute("data-value")
                print("Lookup Code:", lookup_code_value)

                await page.locator("[data-automation-id='tpgServicingTab']").click()

                account_manager = page.locator('[data-automation-id="cboEmployee1"] input')
                await account_manager.wait_for(timeout=10000)
                account_manager_value = await account_manager.input_value()
                print("Account Manager:", account_manager_value)

                await page.locator("//span[normalize-space()='Activities']").click(timeout=5000)
                print("Activities clicked")

                await page.wait_for_timeout(5000)

                access = page.locator("div.main-button", has_text="Access")
                await access.wait_for()
                await access.click()
                print("Access clicked to download attachments")
    
                await page.locator("span.text.force-underline", has_text="Attachments").first.click()
                print("Attachments clicked")

                try:
                    await page.wait_for_selector(
                        "//div[contains(@data-automation-id,'vlvwAttachment body-row')]",
                        state="visible",
                        timeout=10000
                    )
                except:
                    await page.wait_for_timeout(20000)

                rows = page.locator("//div[contains(@data-automation-id,'vlvwAttachment body-row')]")
                row_count = await rows.count()

                print("Total Files:", row_count)

                for i in range(row_count):

                    retry = 0
                    while retry < 3:
                        try:

                            file_xpath = f"(//div[contains(@data-automation-id,'vlvwAttachment body-row')]//div[contains(@class,'body-cell first')]//span[contains(@class,'text')])[{i+1}]"
                            file_element = page.locator(f"xpath={file_xpath}")
                            await file_element.wait_for(state="visible", timeout=15000)

                            file_text = (await file_element.inner_text()).strip()
                            print("Processing:", file_text)

                            await file_element.scroll_into_view_if_needed()
                            await file_element.dblclick()

                            download_btn = page.locator("[data-automation-id='Download']")
                            await download_btn.wait_for(state="visible")

                            async with page.expect_download(timeout=20000) as d:
                                await download_btn.click()

                            download = await d.value

                            file_name = re.sub(r'[<>:"/\\|?*]', '-', download.suggested_filename)

                            save_path = os.path.join(task_folder, file_name)
                            await download.save_as(save_path)

                            print("Saved:", save_path)

                            await page.wait_for_timeout(3000)

                            break

                        except Exception as e:
                            retry += 1
                            print(f"Error at row {i+1} retry {retry}: {e}")
                            await page.wait_for_timeout(3000)


                # =========================
                # FINAL STEP: Process MSG after ALL downloads
                # =========================

                target_keywords = [
                    "quote", "binder", "proposal",
                    "schedule", "application", "app", "accord"
                ]

                existing_types = set()

                # Detect OUTER PDFs
                for f in os.listdir(task_folder):
                    f_lower = f.lower()
                    if f_lower.endswith(".pdf"):
                        for k in target_keywords:
                            if k in f_lower:
                                existing_types.add(k)

                # print("Existing OUTER types:", existing_types)

                # # Process MSG files AFTER
                # for f in os.listdir(task_folder):
                #     if f.lower().endswith(".msg"):
                #         msg_path = os.path.join(task_folder, f)
                #         extract_pdf_from_msg(msg_path, existing_types)

                msg_files = [f for f in os.listdir(task_folder) if f.lower().endswith(".msg")]

                if msg_files:
                    print("Existing OUTER types:", existing_types)

                    for f in msg_files:
                        msg_path = os.path.join(task_folder, f)
                        extract_pdf_from_msg(msg_path, existing_types)


                row_data = {
                    "Date": date_text,
                    "Account Manager": account_manager_value,
                    "Lookup Code": lookup_code_value,
                    "Account Name": account_name_value,
                    "Description": description_value,
                    "Policy number": policy_no,
                    "LOB": lob,
                    "Association": association_value,
                    "Job ID": "",
                    "Status": "Pending",
                    "Reference ID": ref_id
                }

                append_to_excel(row_data, existing_ids)
                existing_ids.add(unique_id)
                save_checkpoint(target_date, row)

                # await page.locator("div.close-button").wait_for(state="visible")
                # await page.locator("div.close-button").click()

                # # close_btn = page.locator("div.close-button[title*='Close']").last
                # # await close_btn.wait_for(state="visible", timeout=10000)
                # # await close_btn.click()

                home_btn = page.locator("div.main-button:has-text('Home')")
                if await home_btn.is_visible():
                    await home_btn.click()
                    print("Clicked home button")

                    await page.get_by_text("Follow Up/Start").wait_for(state="visible", timeout=20000)

                await page.wait_for_timeout(3000)

                print("\n-------------------------------------------------------------------------------------------------------------------------------------------------------------------\n")

                break

            except Exception as e:
                    
                    await page.wait_for_timeout(10000)

                    retry += 1
                    #print(f"Error at row {row} retry {retry}: {e}")
                    print(f"\n ERROR at row {row} retry {retry}")
                    traceback.print_exc()

                    home_btn = page.locator("div.main-button:has-text('Home')")

                    if await home_btn.is_visible():
                        await home_btn.click()
                        print("Clicked Home before retry")

                        await page.get_by_text("Follow Up/Start").wait_for(state="visible", timeout=20000)

                    await page.wait_for_timeout(3000)

                    if retry >= MAX_RETRIES:
                        print(f"Skipping row {row} after {MAX_RETRIES} retries")

                        error_data = {
                            "Date": date_text,
                            "Account Manager": account_manager_value,
                            "Lookup Code": lookup_code_value,
                            "Account Name": account_name_value,
                            "Description": description_value,
                            "Policy number": policy_no,
                            "LOB": lob,
                            "Association": association_value,
                            "Job ID": "",
                            "Status": "Error",
                            "Reference ID": ref_id
                        }

                        append_to_excel(error_data, existing_ids)
                        save_checkpoint(target_date, row)
                        break

        if duplicate_found:
            row += 1
            continue

        row += 1

# SAVE TO EXCEL FUNCTION

def append_to_excel(row_data, existing_ids):

    row_data["Reference ID"] = str(row_data["Reference ID"])

    df = pd.DataFrame([row_data])

    if not os.path.exists(EXCEL_FILE_PATH):
        df.to_excel(EXCEL_FILE_PATH, index=False)
        print(f"Saved → {row_data['Reference ID']}")
        return True

    from openpyxl import load_workbook
    book = load_workbook(EXCEL_FILE_PATH)
    sheet = book.active
    startrow = sheet.max_row
    book.close()

    with pd.ExcelWriter(
        EXCEL_FILE_PATH,
        engine="openpyxl",
        mode="a",
        if_sheet_exists="overlay"
    ) as writer:

        df.to_excel(
            writer,
            index=False,
            header=False,
            startrow=startrow
        )

    print(f"Saved → {row_data['Reference ID']}")
    return True


# MAIN FUNCTION

async def run_epic():

    MAX_BROWSER_RETRIES = 3

    start_id = get_starting_ref_id()

    for attempt in range(MAX_BROWSER_RETRIES):

        browser = None
        context = None
        page = None

        try:
            print(f"\n--------------- EPIC Attempt {attempt+1}")

            existing_ids = set()

            manual_dates = config.get("RUN", "MANUAL_DATES", fallback="").strip()

            if manual_dates:
                target_dates = [
                    datetime.strptime(d.strip(), "%Y-%m-%d").date()
                    for d in manual_dates.split(",")
                    if d.strip()
                ]

                print(f"Manual Run Dates: {target_dates}")

            else:
                today = datetime.now().date()
                weekday = today.weekday()

                if weekday == 0:  # Monday
                    target_dates = [
                        today - timedelta(days=3),  # Friday
                        today - timedelta(days=2),  # Saturday
                        today - timedelta(days=1),  # Sunday
                    ]
                elif weekday in [1, 2, 3, 4]:  # Tuesday-Friday
                    target_dates = [
                        today - timedelta(days=1)
                    ]
                else:
                    target_dates = []

            if os.path.exists(EXCEL_FILE_PATH):
                df = pd.read_excel(EXCEL_FILE_PATH)
                df["Date"] = pd.to_datetime(df["Date"], errors="coerce").dt.date
            else:
                df = pd.DataFrame()

            async with async_playwright() as p:

                browser = await p.chromium.launch(channel="chrome", headless=False)
                context = await browser.new_context()
                page = await context.new_page()

                page.set_default_timeout(60000)
                page.set_default_navigation_timeout(90000)

                await login(page, context)

                for target_date in target_dates:

                    if not df.empty:
                        df_filtered = df[df["Date"] == target_date].copy()

                        existing_ids = set(
                            df_filtered["Association"]
                            .astype(str)
                            .str.strip()
                            .str.lower()
                        )
                    else:
                        existing_ids = set()

                    print("\n----------------------------------------------------------------\n")

                    print(f"Processing Date: {target_date}")

                    await scrape_records(page, start_id, existing_ids, target_date)

                    clear_checkpoint()

                if not page.is_closed():
                    try:
                        await page.get_by_text("Logout").click(timeout=5000)
                        await page.locator("[data-automation-id='Yes']").click(timeout=5000)
                        await page.wait_for_timeout(3000)
                        print("Logout successful")
                    except Exception as e:
                        print("Logout failed:", e)

                print("EPIC completed successfully")
                clear_checkpoint()
                return 

        # except PlaywrightError as e:
        #     error_text = str(e)

        #     print(f"\n Playwright error: {error_text}")

        #     if "Target closed" in error_text:
        #         print("Detected browser crash → restarting...")

        #     await asyncio.sleep(5)
        except PlaywrightError as e:
            print("\n================ PLAYWRIGHT ERROR ================")
            print(type(e).__name__)
            print(str(e))
            traceback.print_exc()     
            print("================================================")

            await asyncio.sleep(5)

        finally:
            print("Cleaning EPIC resources...")

            try:
                if browser and browser.is_connected():
                    await browser.close()
                    print("Browser closed successfully")
            except Exception as e:
                print("Browser cleanup error:", e)

    print("EPIC failed after max retries")


if __name__ == "__main__":
    asyncio.run(run_epic())